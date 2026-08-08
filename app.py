"""
Sales vs Target — Management Dashboard Backend (v2)
====================================================
Business Year (BY) = Oct -> Sep
  Q1: Oct-Dec | Q2: Jan-Mar | Q3: Apr-Jun | Q4: Jul-Sep
  (mapping is driven by the Quarter Master upload)

Six uploaded files:
  1. Customer & Team Master   (Customer Code, Customer Name, Team, MIS Customer)
  2. MIS Category Master      (Product key -> MIS Category)
  3. Quarter Master           (Month number -> BY Quarter)
  4. Target                   (SE/Customer/Product-Third/Sub/Cat + quarterly targets)
  5. Sales                    (SAP export — current BY YTD + entire prior BY)
  6. PO                       (open POs not yet shipped)

Rules:
  - Team for any Sales/Target row is resolved by Customer Code -> Customer Master.
  - MIS Category is resolved by checking Product Third Category, then Product
    Sub Category, then Product Category against the MIS Category map (in that
    sequence) and taking the FIRST hit.
  - Sales values use Sell.Value / 1,000,000 -> USD (Mn.).
  - PO values use Invoice Amt. / 1,000,000 (already in USD per the source file).
"""
import os
import io
import json
import sqlite3
import secrets
import datetime as dt
import logging
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import closing
from functools import wraps

from env_utils import load_env_file

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(name)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

load_env_file(__file__)

import numpy as np
import pandas as pd
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(BASE_DIR, "data"))
if not os.path.isabs(DATA_DIR):
    DATA_DIR = os.path.join(BASE_DIR, DATA_DIR)
DB_PATH = os.path.join(DATA_DIR, "dashboard.db")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")
if DATA_DIR in ("/app/data", "/app/data/"):
    DATA_DIR = os.path.join(BASE_DIR, "data")
    DB_PATH = os.path.join(DATA_DIR, "dashboard.db")
os.makedirs(DATA_DIR, exist_ok=True)

app = Flask(__name__, static_folder=None)
CORS(app)

MONTH_LABELS = ["Oct", "Nov", "Dec", "Jan", "Feb", "Mar",
                "Apr", "May", "Jun", "Jul", "Aug", "Sep"]
QUARTER_LABELS = ["Q1 (Oct-Dec)", "Q2 (Jan-Mar)", "Q3 (Apr-Jun)", "Q4 (Jul-Sep)"]

# Default quarter mapping (calendar month -> BY quarter index 0..3)
# Overridden by the Quarter Master upload if present.
DEFAULT_MONTH_TO_QUARTER = {
    10: 0, 11: 0, 12: 0,   # Q1 Oct-Dec
    1: 1,  2: 1,  3: 1,    # Q2 Jan-Mar
    4: 2,  5: 2,  6: 2,    # Q3 Apr-Jun
    7: 3,  8: 3,  9: 3,    # Q4 Jul-Sep
}


# ──────────────────────────────────────────────────────────────────────
#  DB
# ──────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with closing(get_db()) as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS datasets (
            key        TEXT PRIMARY KEY,
            filename   TEXT,
            uploaded   TEXT,
            rowcount   INTEGER,
            payload    TEXT
        );
        CREATE TABLE IF NOT EXISTS team_note_items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            by          TEXT NOT NULL,
            team        TEXT NOT NULL,
            customer    TEXT NOT NULL DEFAULT '',
            particulars TEXT NOT NULL DEFAULT 'expected',
            value_usd   REAL DEFAULT 0,
            notes       TEXT DEFAULT '',
            status      TEXT DEFAULT 'po_not_sap',
            program     TEXT DEFAULT '',
            period      TEXT DEFAULT '',
            month_key   TEXT DEFAULT '',
            updated_at  TEXT
        );
        CREATE TABLE IF NOT EXISTS users (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            username   TEXT UNIQUE NOT NULL,
            password   TEXT NOT NULL,
            role       TEXT NOT NULL DEFAULT 'team',
            team       TEXT DEFAULT '',
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS tokens (
            token      TEXT PRIMARY KEY,
            user_id    INTEGER,
            username   TEXT,
            role       TEXT,
            team       TEXT DEFAULT '',
            created_at TEXT
        );

        -- ── Market Intelligence ───────────────────────────────────────────
        -- Product categories (manageable via admin UI)
        CREATE TABLE IF NOT EXISTS mi_categories (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT NOT NULL,
            description TEXT DEFAULT '',
            keywords    TEXT DEFAULT '[]',
            is_active   INTEGER DEFAULT 1,
            sort_order  INTEGER DEFAULT 0,
            created_at  TEXT,
            updated_at  TEXT
        );

        -- Products within categories
        CREATE TABLE IF NOT EXISTS mi_products (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            category_id     INTEGER NOT NULL,
            name            TEXT NOT NULL,
            description     TEXT DEFAULT '',
            keywords        TEXT DEFAULT '[]',
            alternate_names TEXT DEFAULT '[]',
            hs_codes        TEXT DEFAULT '[]',
            is_active       INTEGER DEFAULT 1,
            sort_order      INTEGER DEFAULT 0,
            created_at      TEXT,
            updated_at      TEXT
        );

        -- Team → market mapping (manageable via admin UI)
        CREATE TABLE IF NOT EXISTS mi_team_markets (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            team_name    TEXT NOT NULL,
            market       TEXT NOT NULL,
            country_code TEXT DEFAULT '',
            region       TEXT DEFAULT '',
            is_active    INTEGER DEFAULT 1,
            sort_order   INTEGER DEFAULT 0,
            created_at   TEXT
        );

        -- One-shot "run this country/region+category refresh at a specific
        -- date/time" jobs, queued from the Market Information tab.
        CREATE TABLE IF NOT EXISTS mi_oneshot_jobs (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            country      TEXT DEFAULT '',
            region       TEXT DEFAULT '',
            category     TEXT NOT NULL,
            run_at       TEXT NOT NULL,
            status       TEXT NOT NULL DEFAULT 'pending',
            error        TEXT DEFAULT '',
            created_by   TEXT DEFAULT '',
            created_at   TEXT,
            completed_at TEXT
        );

        -- Intelligence updates produced by the monthly job
        CREATE TABLE IF NOT EXISTS mi_updates (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            month             INTEGER NOT NULL,
            year              INTEGER NOT NULL,
            team_name         TEXT NOT NULL,
            market            TEXT NOT NULL,
            category_id       INTEGER,
            category_name     TEXT DEFAULT '',
            product_name      TEXT DEFAULT '',
            update_title      TEXT NOT NULL,
            update_summary    TEXT NOT NULL,
            business_impact   TEXT DEFAULT '',
            source_name       TEXT DEFAULT '',
            source_url        TEXT DEFAULT '',
            source_type       TEXT DEFAULT 'Other',
            publication_date  TEXT DEFAULT '',
            credibility_score REAL DEFAULT 5,
            relevance_score   REAL DEFAULT 5,
            content_hash      TEXT DEFAULT '',
            created_at        TEXT,
            updated_at        TEXT
        );

        -- Job execution log
        CREATE TABLE IF NOT EXISTS mi_job_logs (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            job_type         TEXT NOT NULL DEFAULT 'monthly_fetch',
            status           TEXT NOT NULL DEFAULT 'running',
            month            INTEGER,
            year             INTEGER,
            updates_fetched  INTEGER DEFAULT 0,
            updates_saved    INTEGER DEFAULT 0,
            teams_processed  TEXT DEFAULT '[]',
            error_message    TEXT DEFAULT '',
            started_at       TEXT,
            completed_at     TEXT
        );

        -- ── Market Information Cache ──────────────────────────────────────
        CREATE TABLE IF NOT EXISTS market_info_cache (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            country       TEXT NOT NULL,
            category      TEXT NOT NULL,
            data_json     TEXT NOT NULL DEFAULT '{}',
            fetch_date    TEXT NOT NULL,
            next_update   TEXT NOT NULL,
            attempt_count INTEGER DEFAULT 0,
            cache_status  TEXT DEFAULT 'ok',
            created_at    TEXT,
            updated_at    TEXT,
            UNIQUE(country, category)
        );

        CREATE TABLE IF NOT EXISTS market_info_refresh_log (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            country      TEXT NOT NULL,
            category     TEXT NOT NULL,
            triggered_by TEXT DEFAULT 'manual',
            status       TEXT DEFAULT 'ok',
            error_msg    TEXT DEFAULT '',
            refreshed_at TEXT NOT NULL
        );

        -- Per-user tab access control
        CREATE TABLE IF NOT EXISTS tab_access (
            tab_name  TEXT NOT NULL,
            user_id   INTEGER NOT NULL,
            PRIMARY KEY (tab_name, user_id)
        );

        -- Monthly sales-review meeting minutes
        CREATE TABLE IF NOT EXISTS meeting_minutes (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            by           TEXT NOT NULL,
            month_key    TEXT NOT NULL,
            meeting_date TEXT NOT NULL,
            title        TEXT DEFAULT '',
            attendees    TEXT DEFAULT '',
            circulated_to TEXT DEFAULT '',
            notes        TEXT DEFAULT '',
            created_by   TEXT DEFAULT '',
            created_at   TEXT,
            updated_at   TEXT
        );

        -- Action points arising from meeting minutes
        CREATE TABLE IF NOT EXISTS meeting_action_items (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            minute_id     INTEGER NOT NULL,
            description   TEXT NOT NULL DEFAULT '',
            responsible   TEXT DEFAULT '',
            deadline      TEXT DEFAULT '',
            status        TEXT DEFAULT 'open',
            created_at    TEXT,
            updated_at    TEXT
        );

        -- Personal/team meeting notes recorded by team members
        CREATE TABLE IF NOT EXISTS team_meeting_notes (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            team         TEXT NOT NULL,
            note_date    TEXT NOT NULL,
            title        TEXT DEFAULT '',
            notes        TEXT DEFAULT '',
            created_by   TEXT DEFAULT '',
            created_at   TEXT,
            updated_at   TEXT
        );
        """)
    # Migrations for existing installs
    with closing(get_db()) as db:
        try:
            db.execute("ALTER TABLE team_note_items ADD COLUMN status TEXT DEFAULT 'po_not_sap'")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE mi_categories ADD COLUMN hs_codes TEXT DEFAULT '[]'")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE mi_team_markets ADD COLUMN region TEXT DEFAULT ''")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE team_note_items ADD COLUMN program TEXT DEFAULT ''")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE team_note_items ADD COLUMN period TEXT DEFAULT ''")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE team_note_items ADD COLUMN month_key TEXT DEFAULT ''")
            db.commit()
        except Exception:
            pass
        try:
            # Backfill month_key for pre-existing rows so they don't vanish from the
            # "current month" view; treat them as belonging to the current month.
            db.execute(
                "UPDATE team_note_items SET month_key=? WHERE month_key IS NULL OR month_key=''",
                (dt.date.today().strftime("%Y-%m"),)
            )
            db.commit()
        except Exception:
            pass
    # Always ensure admin exists with current ADMIN_PASSWORD
    with closing(get_db()) as db:
        db.execute(
            """INSERT INTO users (username,password,role,team,created_at)
               VALUES (?,?,'admin','',?)
               ON CONFLICT(username) DO UPDATE SET password=excluded.password,role='admin'""",
            ("admin", generate_password_hash(ADMIN_PASSWORD),
             dt.datetime.now().isoformat(timespec="seconds"))
        )
        db.commit()
    # One-time: grant the admin account full tab access by default (toggleable afterwards)
    with closing(get_db()) as db:
        try:
            seeded = db.execute("SELECT 1 FROM datasets WHERE key='_admin_tab_access_seeded'").fetchone()
            if not seeded:
                admin_row = db.execute("SELECT id FROM users WHERE username='admin'").fetchone()
                if admin_row:
                    for tab in ("pm", "mn", "mi"):
                        db.execute("INSERT OR IGNORE INTO tab_access (tab_name,user_id) VALUES (?,?)",
                                   (tab, admin_row["id"]))
                db.execute(
                    "INSERT OR REPLACE INTO datasets (key,filename,uploaded,rowcount,payload) "
                    "VALUES ('_admin_tab_access_seeded','','',0,'1')"
                )
                db.commit()
        except Exception:
            pass


def _seed_mi_defaults():
    """Seed default product categories, products, and team-market mappings
    into MI tables on first run.  Safe to call repeatedly — checks if empty first.
    To add more categories/products/teams: use the Admin UI instead of editing here."""
    DEFAULT_CATEGORIES = [
        {
            "name": "Paper Stationery",
            "description": "Paper-based stationery — notebooks, exercise books, file folders",
            "keywords": ["paper", "stationery", "notebook", "exercise book", "journal", "file folder"],
            "sort_order": 1,
            "products": [
                {"name": "Notebooks",      "keywords": ["notebook","spiral notebook","composition book"], "alternate_names": ["notepad","writing pad"]},
                {"name": "Exercise Books", "keywords": ["exercise book","school book","composition"],     "alternate_names": ["school notebook"]},
                {"name": "Journals",       "keywords": ["journal","diary","planner"],                    "alternate_names": ["diary"]},
                {"name": "File Folders",   "keywords": ["file folder","document folder","manila folder"],"alternate_names": ["folder","binder"]},
            ],
        },
        {
            "name": "Injection Molding Products",
            "description": "Plastic office and stationery products",
            "keywords": ["plastic","injection molding","desk organizer","office storage"],
            "sort_order": 2,
            "products": [
                {"name": "Desk Organizers",           "keywords": ["desk organizer","desktop organizer","pen holder"],    "alternate_names": ["desk caddy","pencil holder"]},
                {"name": "Plastic Office Organizers", "keywords": ["plastic organizer","office organizer","document tray"],"alternate_names": ["file tray"]},
                {"name": "Plastic Storage Products",  "keywords": ["plastic storage","storage box","storage tray"],       "alternate_names": ["storage container"]},
            ],
        },
        {
            "name": "Writing and Correction Products",
            "description": "Writing instruments and correction products",
            "keywords": ["pencil","eraser","sharpener","writing","correction"],
            "sort_order": 3,
            "products": [
                {"name": "Pencils",    "keywords": ["pencil","HB pencil","graphite pencil","colored pencil"],"alternate_names": ["lead pencil"]},
                {"name": "Erasers",    "keywords": ["eraser","rubber eraser","pencil eraser"],               "alternate_names": ["rubber"]},
                {"name": "Sharpeners", "keywords": ["pencil sharpener","sharpener"],                        "alternate_names": []},
            ],
        },
        {
            "name": "Metal Stationery Products",
            "description": "Metal-based stationery — staplers, clips, staples",
            "keywords": ["metal","stapler","paper clip","binder clip","staples"],
            "sort_order": 4,
            "products": [
                {"name": "Paper Clips",  "keywords": ["paper clip","paperclip","gem clip"],              "alternate_names": ["gem clip"]},
                {"name": "Binder Clips", "keywords": ["binder clip","bulldog clip","fold-back clip"],    "alternate_names": ["bulldog clip"]},
                {"name": "Staplers",     "keywords": ["stapler","desk stapler","heavy duty stapler"],    "alternate_names": []},
                {"name": "Staples",      "keywords": ["staple","staple pin","stapler pin","staple refill"],"alternate_names": ["staple pins","stapler refill"]},
            ],
        },
    ]
    DEFAULT_TEAM_MARKETS = [
        # Add your sales team → market mappings here (or via Admin UI later)
        {"team_name": "USA 1",   "market": "USA",         "country_code": "US", "sort_order": 1},
        {"team_name": "USA 1",   "market": "Canada",      "country_code": "CA", "sort_order": 2},
        {"team_name": "USA 1",   "market": "Mexico",      "country_code": "MX", "sort_order": 3},
        {"team_name": "Oceanic", "market": "Australia",   "country_code": "AU", "sort_order": 1},
        {"team_name": "Oceanic", "market": "New Zealand", "country_code": "NZ", "sort_order": 2},
    ]
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        if db.execute("SELECT 1 FROM mi_categories LIMIT 1").fetchone():
            return  # already seeded
        for i, cat in enumerate(DEFAULT_CATEGORIES):
            db.execute(
                "INSERT INTO mi_categories (name,description,keywords,is_active,sort_order,created_at,updated_at) VALUES(?,?,?,1,?,?,?)",
                (cat["name"], cat["description"], json.dumps(cat["keywords"]), cat.get("sort_order", i), now, now)
            )
            cat_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
            for j, prod in enumerate(cat["products"]):
                db.execute(
                    "INSERT INTO mi_products (category_id,name,keywords,alternate_names,hs_codes,is_active,sort_order,created_at,updated_at) VALUES(?,?,?,?,?,1,?,?,?)",
                    (cat_id, prod["name"], json.dumps(prod.get("keywords", [])),
                     json.dumps(prod.get("alternate_names", [])), json.dumps([]), j, now, now)
                )
        for tm in DEFAULT_TEAM_MARKETS:
            db.execute(
                "INSERT INTO mi_team_markets (team_name,market,country_code,is_active,sort_order,created_at) VALUES(?,?,?,1,?,?)",
                (tm["team_name"], tm["market"], tm["country_code"], tm["sort_order"], now)
            )
        db.commit()
        logger.info("MI defaults seeded")


def save_dataset(key, filename, records):
    with closing(get_db()) as db:
        db.execute(
            "REPLACE INTO datasets (key, filename, uploaded, rowcount, payload) "
            "VALUES (?,?,?,?,?)",
            (key, filename, dt.datetime.now().isoformat(timespec="seconds"),
             len(records), json.dumps(records)),
        )
        db.commit()


def load_dataset(key):
    with closing(get_db()) as db:
        row = db.execute(
            "SELECT filename, uploaded, rowcount, payload FROM datasets WHERE key=?",
            (key,)).fetchone()
        if not row:
            return None
        return {
            "filename": row["filename"],
            "uploaded": row["uploaded"],
            "rowcount": row["rowcount"],
            "records": json.loads(row["payload"]),
        }


def dataset_meta(key):
    with closing(get_db()) as db:
        row = db.execute(
            "SELECT filename, uploaded, rowcount FROM datasets WHERE key=?",
            (key,)).fetchone()
        if not row:
            return None
        return {"filename": row["filename"], "uploaded": row["uploaded"],
                "rowcount": row["rowcount"]}


def clear_key(key):
    with closing(get_db()) as db:
        db.execute("DELETE FROM datasets WHERE key=?", (key,))
        db.commit()


# ──────────────────────────────────────────────────────────────────────
#  Parsing helpers
# ──────────────────────────────────────────────────────────────────────
def read_excel_any(file_storage):
    """Read xlsx/xls/xlsb/csv into a dict of {sheet_name: DataFrame(header=None)}."""
    name = file_storage.filename.lower()
    raw = file_storage.read()
    bio = io.BytesIO(raw)
    if name.endswith(".csv"):
        df = pd.read_csv(bio, header=None, dtype=object)
        return {"Sheet1": df}
    engine = None
    if name.endswith(".xlsb"):
        engine = "pyxlsb"
    elif name.endswith(".xls"):
        engine = "xlrd"
    else:
        engine = "openpyxl"
    xls = pd.read_excel(bio, sheet_name=None, header=None, engine=engine, dtype=object)
    return xls


def norm(s):
    return str(s).strip().lower().replace(" ", "").replace("_", "").replace(".", "").replace("-", "").replace("/", "")


def find_header_row(df, needles, scan=15):
    needles_n = [norm(n) for n in needles]
    best_i, best_hits = 0, 0
    for i in range(min(scan, len(df))):
        row = [norm(x) for x in df.iloc[i].tolist()]
        hits = sum(1 for nd in needles_n if any(nd in c for c in row if c))
        if hits > best_hits:
            best_hits, best_i = hits, i
    if best_hits >= max(2, len(needles) - 2):
        return best_i
    return 0


def df_with_header(df, needles):
    hr = find_header_row(df, needles)
    header = [str(x).strip() if x is not None else "" for x in df.iloc[hr].tolist()]
    body = df.iloc[hr + 1:].copy()
    body.columns = header
    body = body.loc[:, [c for c in body.columns if c]]
    # Drop duplicate columns by keeping first
    body = body.loc[:, ~pd.Index(body.columns).duplicated()]
    body = body.dropna(how="all")
    return body


def col(row, *needles):
    """Fuzzy column getter for a pandas Series row."""
    keys = list(row.index)
    kn = {k: norm(k) for k in keys}
    for nd in needles:
        ndn = norm(nd)
        for k in keys:
            if kn[k] == ndn:
                v = row[k]
                if v is not None and str(v).strip() != "" and not (isinstance(v, float) and np.isnan(v)):
                    return v
    for nd in needles:
        ndn = norm(nd)
        for k in keys:
            if ndn in kn[k]:
                v = row[k]
                if v is not None and str(v).strip() != "" and not (isinstance(v, float) and np.isnan(v)):
                    return v
    return None


def to_num(v):
    try:
        if v is None or str(v).strip() == "":
            return 0.0
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def excel_date(v):
    """Convert excel serial / string / datetime to date object."""
    if v is None or v == "":
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v if isinstance(v, dt.date) else v.date()
    if isinstance(v, (int, float)) and not (isinstance(v, float) and np.isnan(v)):
        try:
            return (dt.datetime(1899, 12, 30) + dt.timedelta(days=float(v))).date()
        except (ValueError, OverflowError):
            return None
    try:
        return pd.to_datetime(v, dayfirst=True).date()
    except (ValueError, TypeError):
        try:
            return pd.to_datetime(v).date()
        except (ValueError, TypeError):
            return None


def business_year(d):
    """Return BY string e.g. '2025-26' for Oct2025..Sep2026."""
    if d is None:
        return None
    if d.month >= 10:
        return f"{d.year}-{str((d.year + 1) % 100).zfill(2)}"
    return f"{d.year - 1}-{str(d.year % 100).zfill(2)}"


def by_month_index(d):
    """0=Oct ... 11=Sep"""
    return (d.month + 2) % 12 if d.month >= 10 else d.month + 2


def cust_code_norm(v):
    """Normalise customer code: strip whitespace, drop trailing .0 from floats."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def get_quarter_map():
    """Return dict {calendar month (1..12): quarter index 0..3} from
    the uploaded Quarter Master, or the default Oct->Sep mapping."""
    qm_ds = load_dataset("quarter_master")
    if not qm_ds:
        return dict(DEFAULT_MONTH_TO_QUARTER)
    out = {}
    for rec in qm_ds["records"]:
        m = rec.get("month")
        q = rec.get("by_q")
        if m and q:
            try:
                out[int(m)] = int(q) - 1   # store 0..3
            except (ValueError, TypeError):
                continue
    if len(out) >= 12:
        return out
    # Fallback if incomplete
    merged = dict(DEFAULT_MONTH_TO_QUARTER)
    merged.update(out)
    return merged


def get_mis_map():
    """Return dict {normalised product key -> MIS Category}."""
    mc_ds = load_dataset("mis_category")
    if not mc_ds:
        return {}
    return {norm(r["key"]): r["mis"] for r in mc_ds["records"] if r.get("key") and r.get("mis")}


def get_cust_map():
    """Return dict {customer_code -> {name, team, mis_customer}}."""
    cm_ds = load_dataset("cust_team_master")
    if not cm_ds:
        return {}
    return {cust_code_norm(r["code"]): r for r in cm_ds["records"]}


def resolve_mis_category(third, sub, cat, mis_map):
    """Priority: Product Third Category > Product Sub Category > Product Category."""
    for key in (third, sub, cat):
        if key and str(key).strip():
            v = mis_map.get(norm(key))
            if v:
                return v
    return None


def resolve_team(cust_code, cust_map):
    """Team comes ONLY from the Customer & Team Master via Customer Code."""
    rec = cust_map.get(cust_code_norm(cust_code))
    return rec["team"] if rec else None


def resolve_mis_customer(cust_code, fallback_name, cust_map):
    rec = cust_map.get(cust_code_norm(cust_code))
    if rec and rec.get("mis_customer"):
        return rec["mis_customer"]
    if rec and rec.get("name"):
        return rec["name"]
    return fallback_name or "—"


# ──────────────────────────────────────────────────────────────────────
#  File-specific parsers
# ──────────────────────────────────────────────────────────────────────
def parse_cust_team_master(sheets):
    df = list(sheets.values())[0]
    body = df_with_header(df, ["Customer Code", "Customer Name", "Team", "MIS Customer"])
    out = []
    for _, r in body.iterrows():
        code = col(r, "Customer Code", "CustCode", "Code")
        if code is None:
            continue
        out.append({
            "code": cust_code_norm(code),
            "name": str(col(r, "Customer Name", "Name") or "").strip(),
            "team": str(col(r, "Team") or "").strip(),
            "mis_customer": str(col(r, "MIS Customer", "MIS") or "").strip(),
            "region": str(col(r, "Region", "Export Region") or "").strip(),
        })
    return out


def parse_mis_category(sheets):
    """MIS Category sheet: a 'key' column listing Product Third/Sub/Cat names
    and an 'MIS Category' column with the consolidated label."""
    df = list(sheets.values())[0]
    body = df_with_header(df, ["MIS Category"])
    # Find key column (header may be a long sentence like
    # 'Product Third Category / Product Sub Category / Product Category')
    cols = list(body.columns)
    mis_col = None
    for c in cols:
        if norm(c) == norm("MIS Category"):
            mis_col = c
            break
    if mis_col is None:
        for c in cols:
            if "mis" in norm(c):
                mis_col = c
                break
    key_col = None
    for c in cols:
        if c == mis_col:
            continue
        cn = norm(c)
        if "product" in cn or "category" in cn or "subcategory" in cn or "third" in cn:
            key_col = c
            break
    if key_col is None:
        # First non-empty / non-MIS column
        for c in cols:
            if c != mis_col and c.strip():
                key_col = c
                break

    out = []
    if not mis_col or not key_col:
        return out
    for _, r in body.iterrows():
        k = r.get(key_col)
        v = r.get(mis_col)
        if k is None or v is None:
            continue
        ks = str(k).strip()
        vs = str(v).strip()
        if not ks or not vs or ks.lower() == "nan" or vs.lower() == "nan":
            continue
        out.append({"key": ks, "mis": vs})
    return out


def parse_quarter_master(sheets):
    df = list(sheets.values())[0]
    body = df_with_header(df, ["Month", "BY"])
    out = []
    for _, r in body.iterrows():
        m = col(r, "Month")
        q = col(r, "BY Qrt", "BY Quarter", "BY Qtr", "BYQrt")
        if m is None or q is None:
            continue
        try:
            mi = int(float(m))
            qs = str(q).strip().upper().replace("Q", "")
            qi = int(qs)
        except (ValueError, TypeError):
            continue
        if 1 <= mi <= 12 and 1 <= qi <= 4:
            out.append({"month": mi, "by_q": qi,
                        "mmm": str(col(r, "MMM") or "")})
    return out


def parse_target(sheets):
    names = list(sheets.keys())
    exact = [n for n in names if n.strip().lower() == "target"]
    loose = [n for n in names if "target" in n.lower()]
    name = exact[0] if exact else (loose[0] if loose else names[0])
    df = sheets[name]
    body = df_with_header(df, ["Customer Code", "OCT-DEC Target", "JAN-MAR Target"])
    out = []
    for _, r in body.iterrows():
        code = col(r, "Customer Code")
        if code is None:
            continue
        q = [
            to_num(col(r, "OCT-DEC Target", "Oct-Dec Target", "Q1 Target")),
            to_num(col(r, "JAN-MAR Target", "Jan-Mar Target", "Q2 Target")),
            to_num(col(r, "APR-JUN Target", "Apr-Jun Target", "Q3 Target")),
            to_num(col(r, "JUL-SEP Target", "Jul-Sep Target", "Q4 Target")),
        ]
        if all(x == 0 for x in q):
            continue
        out.append({
            "se_code": str(col(r, "SE Code") or "").strip(),
            "se_name": str(col(r, "SE Name") or "").strip(),
            "code": cust_code_norm(code),
            "name": str(col(r, "Customer Name") or "").strip(),
            "third": str(col(r, "Product Third Category") or "").strip(),
            "category": str(col(r, "Product Category") or "").strip(),
            "sub_category": str(col(r, "Product Sub Category") or "").strip(),
            "targets": q,
        })
    return out


def parse_sales(sheets):
    names = list(sheets.keys())
    exact = [n for n in names if n.strip().lower() == "sales"]
    loose = [n for n in names if "sales" in n.lower()]
    name = exact[0] if exact else (loose[0] if loose else names[0])
    df = sheets[name]
    body = df_with_header(df, ["Customer Code", "Sell.Value", "SAP Invoice Date"])
    out = []
    for _, r in body.iterrows():
        d = excel_date(col(r, "SAP Invoice Date", "Invoice Date", "Inv Date"))
        if d is None:
            continue
        sell = to_num(col(r, "Sell.Value", "Sell Value", "SellValue"))
        if sell == 0:
            continue
        out.append({
            "y": d.year, "m": d.month, "d": d.day,
            "code": cust_code_norm(col(r, "Customer Code")),
            "third": str(col(r, "Product Third Category") or "").strip(),
            "sub_category": str(col(r, "Product Sub Category") or "").strip(),
            "category": str(col(r, "Product Category") or "").strip(),
            "country": str(col(r, "Country of Destination",
                                "Consignee Destination Country",
                                "Destination Country") or "").strip(),
            "region": str(col(r, "Export Region", "Region") or "").strip(),
            "export_id": str(col(r, "Export ID", "Export Id", "ExportID") or "").strip(),
            "sell_mn": sell / 1_000_000.0,
        })
    return out


def parse_po(sheets):
    """PO file uses the same layout as Sales but value is in 'Invoice Amt.' (USD).
    SAP Invoice Date here is the expected ship date (future)."""
    names = list(sheets.keys())
    name = names[0]
    df = sheets[name]
    body = df_with_header(df, ["Customer Code", "Invoice Amt", "PO Number"])
    out = []
    for _, r in body.iterrows():
        # value: prefer Invoice Amt., else Sell.Value
        amt = to_num(col(r, "Invoice Amt.", "Invoice Amt", "InvoiceAmt"))
        if amt == 0:
            amt = to_num(col(r, "Sell.Value", "Sell Value"))
        if amt == 0:
            continue
        d = excel_date(col(r, "SAP Invoice Date", "Expected Ship Date"))
        po_date = excel_date(col(r, "PO Date"))
        out.append({
            "y": d.year if d else None, "m": d.month if d else None, "d": d.day if d else None,
            "po_y": po_date.year if po_date else None,
            "po_m": po_date.month if po_date else None,
            "po_d": po_date.day if po_date else None,
            "code": cust_code_norm(col(r, "Customer Code")),
            "third": str(col(r, "Product Third Category") or "").strip(),
            "sub_category": str(col(r, "Product Sub Category") or "").strip(),
            "category": str(col(r, "Product Category") or "").strip(),
            "country": str(col(r, "Country of Destination",
                                "Consignee Destination Country") or "").strip(),
            "region": str(col(r, "Export Region", "Region") or "").strip(),
            "po_number": str(col(r, "PO Number") or "").strip(),
            "export_id": str(col(r, "Export ID", "Export Id", "ExportID") or "").strip(),
            "amt_mn": amt / 1_000_000.0,
        })
    return out


def parse_prod_cat_master(sheets):
    """Product Code → category hierarchy for MIS category resolution."""
    df = list(sheets.values())[0]
    body = df_with_header(df, ["MATERIAL", "Product Category"])
    out = []
    for _, r in body.iterrows():
        mat = col(r, "MATERIAL", "Material")
        if mat is None:
            continue
        mat_s = str(mat).strip().lstrip("0") or "0"
        third = str(col(r, "Final Category Code Description",
                         "Final Category Description", "Final Category") or "").strip()
        sub = str(col(r, "Sub Product Category Code Description",
                       "Sub Product Category Description",
                       "Sub Product Category") or "").strip()
        cat = str(col(r, "Product Category Code Description",
                       "Product Category Description",
                       "Product Category") or "").strip()
        if not mat_s:
            continue
        out.append({"material": mat_s, "third": third, "sub": sub, "cat": cat})
    return out


def parse_zexpr24(sheets):
    """Program master: Export ID → customer, material, validity dates, FOB price."""
    df = list(sheets.values())[0]
    body = df_with_header(df, ["Export ID", "From date"])
    out = []
    for _, r in body.iterrows():
        eid = col(r, "Export ID", "Export Id")
        mat = col(r, "Material")
        if not eid:
            continue
        from_d = excel_date(col(r, "From date", "From Date"))
        to_d = excel_date(col(r, "To date", "To Date"))
        cust = col(r, "Customer")
        out.append({
            "export_id": str(eid).strip(),
            "material": str(mat).strip().lstrip("0") if mat else "",
            "customer_code": cust_code_norm(cust) if cust else "",
            "customer_name": str(col(r, "Customer Name") or "").strip(),
            "fob_price": to_num(col(r, "FOB Price", "Fob Price")),
            "from_y": from_d.year if from_d else None,
            "from_m": from_d.month if from_d else None,
            "from_d_day": from_d.day if from_d else None,
            "to_y": to_d.year if to_d else None,
            "to_m": to_d.month if to_d else None,
            "to_d_day": to_d.day if to_d else None,
        })
    return out


def _parse_num_cost(v):
    """Parse SAP numeric cell: handles commas, trailing minus, NaN."""
    s = str(v).strip().replace(",", "").replace(" ", "")
    if not s or s.lower() == "nan":
        return 0.0
    neg = s.endswith("-")
    if neg:
        s = s[:-1]
    try:
        val = float(s)
        return -val if neg else val
    except ValueError:
        return 0.0


def parse_zecost06_excel(sheets):
    """Parse zecost06 when it is a real Excel workbook."""
    df = list(sheets.values())[0]
    body = df_with_header(df, ["Customer code", "Export Id", "Material"])
    out = []
    for _, r in body.iterrows():
        export_id = str(col(r, "Export Id", "Export ID") or "").strip()
        material = str(col(r, "Material") or "").strip().lstrip("0") or "0"
        if not export_id or not material or export_id.lower() == "nan":
            continue
        out.append({
            "customer_code": cust_code_norm(col(r, "Customer code", "Customer Code", "Cust code", "Cust Code")),
            "export_id": export_id,
            "material": material,
            "po_qty":        _parse_num_cost(col(r, "PO Qty", "PO qty", "PO Quantity") or 0),
            "po_cost_usd":   _parse_num_cost(col(r, "PO Cost(USD)", "PO Cost (USD)", "PO Cost USD", "PO cost(USD)", "PO Cost") or 0),
            "po_sale_usd":   _parse_num_cost(col(r, "Po Sale Val(USD)", "PO Sale Val(USD)", "PO Sale Val (USD)", "PO Sale Value(USD)", "PO Sale Value (USD)", "PO Sale Value USD", "PO Sale Val", "PO Sale Value") or 0),
            "quot_qty":      _parse_num_cost(col(r, "Quotation", "Quot Qty", "Quot qty", "Quotation Qty", "Quot Quantity") or 0),
            "quot_cost_usd": _parse_num_cost(col(r, "Quot.Cost(USD)", "Quot. Cost(USD)", "Quot Cost(USD)", "Quot. Cost (USD)", "Quot Cost (USD)", "Quot Cost USD", "Quotation Cost(USD)", "Quotation Cost (USD)", "Quot Cost") or 0),
            "quot_sale_usd": _parse_num_cost(col(r, "Quot. Sale Val (USD)", "Quot.Sale Val(USD)", "Quot Sale Val(USD)", "Quot. Sale Value (USD)", "Quot Sale Value(USD)", "Quot Sale Value (USD)", "Quot Sale Value USD", "Quotation Sale Val(USD)", "Quotation Sale Value (USD)", "Quot Sale Val") or 0),
            # Category columns present in the SAP export — use directly to avoid lookup mismatches
            "mis_cat": str(col(r, "MIS Category") or "").strip(),
            "third":   str(col(r, "pod cat level 3", "Product Third Category", "Third Category") or "").strip(),
            "sub":     str(col(r, "pod cat level 2", "Product Sub Category", "Sub Category") or "").strip(),
            "cat":     str(col(r, "pod cat level 1", "Product Category") or "").strip(),
        })
    return out


def parse_zecost06(raw, filename):
    """Parse zecost06 cost report — handles real .xlsx and pipe-delimited text."""
    # Try Excel first
    try:
        bio = io.BytesIO(raw)
        name = filename.lower()
        if name.endswith(".xlsb"):
            engine = "pyxlsb"
        elif name.endswith(".xls"):
            engine = "xlrd"
        else:
            engine = "openpyxl"
        xls = pd.read_excel(bio, sheet_name=None, header=None, engine=engine, dtype=object)
        result = parse_zecost06_excel(xls)
        if result:
            return result
    except Exception:
        pass

    # Fall back to pipe-delimited text (original .xls-as-text format)
    try:
        text = raw.decode("utf-8", errors="replace")
    except Exception:
        text = raw.decode("latin-1", errors="replace")

    lines = text.splitlines()
    cols = []
    header_i = None
    for i, line in enumerate(lines):
        if "|" in line and "Customer code" in line and "Export Id" in line:
            parts = line.split("|")
            cols = [p.strip() for p in parts[1:-1]]
            header_i = i
            break

    if header_i is None:
        return []

    out = []
    for line in lines[header_i + 1:]:
        if not line.startswith("|"):
            continue
        parts = line.split("|")
        values = [p.strip() for p in parts[1:-1]]
        if len(values) < len(cols):
            continue
        if not values[0] or values[0].startswith("-"):
            continue
        row = {cols[i]: values[i] for i in range(min(len(cols), len(values)))}
        export_id = row.get("Export Id", "").strip()
        material = row.get("Material", "").strip().lstrip("0") or "0"
        if not export_id or not material:
            continue
        out.append({
            "customer_code": cust_code_norm(row.get("Customer code", "")),
            "export_id": export_id,
            "material": material,
            "po_qty":       _parse_num_cost(row.get("PO Qty", 0)),
            "po_cost_usd":  _parse_num_cost(row.get("PO Cost(USD)", 0)),
            "po_sale_usd":  _parse_num_cost(row.get("Po Sale Val(USD)", 0)),
            "quot_qty":     _parse_num_cost(row.get("Quotation", 0)),
            "quot_cost_usd":_parse_num_cost(row.get("Quot.Cost(USD)", 0)),
            "quot_sale_usd":_parse_num_cost(row.get("Quot. Sale Val (USD)", 0)),
        })
    return out


PARSERS = {
    "cust_team_master": (parse_cust_team_master, ["Customer Code", "Team"]),
    "mis_category":     (parse_mis_category,     ["MIS Category"]),
    "quarter_master":   (parse_quarter_master,   ["Month", "BY"]),
    "target":           (parse_target,           ["Customer Code", "Target"]),
    "sales":            (parse_sales,            ["Sell.Value", "Invoice"]),
    "po":               (parse_po,               ["PO Number", "Invoice Amt"]),
    "zexpr24":          (parse_zexpr24,          ["Export ID", "From date"]),
    "prod_cat_master":  (parse_prod_cat_master,  ["MATERIAL"]),
    "zecost06":         (parse_zecost06,          ["Cost Analysis"]),
}
TEXT_PARSERS = {"zecost06"}   # receive (raw_bytes, filename) instead of sheets


# ──────────────────────────────────────────────────────────────────────
#  Analytics
# ──────────────────────────────────────────────────────────────────────
def build_enriched():
    """Join sales with masters; returns list of enriched dicts."""
    sales = load_dataset("sales")
    if not sales:
        return None
    cust_map = get_cust_map()
    mis_map = get_mis_map()
    qmap = get_quarter_map()

    enriched = []
    for s in sales["records"]:
        d = dt.date(s["y"], s["m"], s["d"])
        by = business_year(d)
        midx = by_month_index(d)
        q = qmap.get(s["m"], midx // 3)

        cm = cust_map.get(s["code"], {})
        team = cm.get("team") or "—"
        mis_cust = cm.get("mis_customer") or cm.get("name") or "—"
        mis_cat = resolve_mis_category(s.get("third"), s.get("sub_category"),
                                       s.get("category"), mis_map) or "Uncategorised"
        enriched.append({
            "by": by, "midx": midx, "q": q,
            "team": team, "mis_customer": mis_cust, "mis_cat": mis_cat,
            "third": s.get("third") or "—",
            "country": s.get("country") or "—",
            "region": cm.get("region") or "—",
            "sell_mn": s["sell_mn"],
        })
    return enriched


def build_po_enriched():
    """Enrich PO records similarly. We keep all of them (open POs) and tag
    them with the BY they are expected to ship in."""
    po = load_dataset("po")
    if not po:
        return []
    cust_map = get_cust_map()
    mis_map = get_mis_map()
    qmap = get_quarter_map()

    out = []
    for p in po["records"]:
        ship_d = dt.date(p["y"], p["m"], p["d"]) if p.get("y") else None
        po_d = dt.date(p["po_y"], p["po_m"], p["po_d"]) if p.get("po_y") else None
        by = business_year(ship_d) if ship_d else None
        midx = by_month_index(ship_d) if ship_d else None
        q = qmap.get(p["m"], (midx // 3) if midx is not None else None) if p.get("m") else None

        cm = cust_map.get(p["code"], {})
        team = cm.get("team") or "—"
        mis_cust = cm.get("mis_customer") or cm.get("name") or "—"
        mis_cat = resolve_mis_category(p.get("third"), p.get("sub_category"),
                                       p.get("category"), mis_map) or "Uncategorised"
        out.append({
            "by": by, "midx": midx, "q": q,
            "ship_date": ship_d.isoformat() if ship_d else None,
            "po_date": po_d.isoformat() if po_d else None,
            "team": team, "mis_customer": mis_cust, "mis_cat": mis_cat,
            "third": p.get("third") or "—",
            "country": p.get("country") or "—",
            "region": cm.get("region") or "—",
            "po_number": p.get("po_number") or "",
            "amt_mn": p["amt_mn"],
        })
    return out


def parse_teams(team_str):
    """Split comma-separated team string into a set; empty string → empty set."""
    if not team_str:
        return set()
    return {t.strip() for t in str(team_str).split(",") if t.strip()}


# Fixed display order for team dropdowns/lists across the app.
TEAM_SORT_ORDER = [
    "USA 1", "USA 2", "USA 3", "USA 4", "USA 5",
    "Americas", "Europe", "Oceanic",
    "Africa 1", "ME1", "Africa 2", "ME2", "Ecom",
]
_TEAM_SORT_INDEX = {t: i for i, t in enumerate(TEAM_SORT_ORDER)}


def sort_teams(teams):
    """Sort an iterable of team names by the fixed TEAM_SORT_ORDER, with any
    unlisted teams appended afterwards in alphabetical order."""
    teams = list(teams)
    return sorted(teams, key=lambda t: (_TEAM_SORT_INDEX.get(t, len(TEAM_SORT_ORDER)), t))


def resolve_team_filter(u, requested_team):
    """For team-role users, restrict the requested team filter to their assigned
    teams. If the request isn't a subset of what they're allowed to see, fall back
    to all of their assigned teams."""
    if not (u and u["role"] == "team" and u["team"]):
        return requested_team
    allowed = parse_teams(u["team"])
    requested = parse_teams(requested_team or "")
    if requested and requested.issubset(allowed):
        return requested_team
    return u["team"]


def prior_by(by):
    a, b = by.split("-")
    return f"{int(a) - 1}-{str(int(a) % 100).zfill(2)}"


def target_ytd(qtargets, up_to_month):
    """Proportional YTD target: full completed quarters + linear partial quarter."""
    last_q = up_to_month // 3
    months_into_q = (up_to_month % 3) + 1
    total = 0.0
    for q in range(4):
        v = qtargets[q] / 1_000_000.0
        if q < last_q:
            total += v
        elif q == last_q:
            total += v * (months_into_q / 3.0)
    return total


def notes_overall_summary(by, f_team):
    """Aggregate Ongoing-negotiations (current month, by status) and Business-Lost
    totals across the given set of teams (or all teams if f_team is empty), for
    the given Business Year. Used by the Overall Business / Executive Summary."""
    cur_month = _current_month_key()
    with closing(get_db()) as db:
        q = ("SELECT status, SUM(value_usd) AS total FROM team_note_items "
             "WHERE by=? AND particulars='negotiation' AND month_key=?")
        params = [by, cur_month]
        if f_team:
            placeholders = ",".join("?" * len(f_team))
            q += f" AND team IN ({placeholders})"
            params.extend(sorted(f_team))
        q += " GROUP BY status"
        neg_rows = db.execute(q, params).fetchall()
        neg_by_status = {r["status"]: round(r["total"] or 0.0, 3) for r in neg_rows}

        q2 = "SELECT SUM(value_usd) AS total FROM team_note_items WHERE by=? AND particulars='business_lost'"
        params2 = [by]
        if f_team:
            placeholders = ",".join("?" * len(f_team))
            q2 += f" AND team IN ({placeholders})"
            params2.extend(sorted(f_team))
        bl_row = db.execute(q2, params2).fetchone()
        business_lost_total = round((bl_row["total"] if bl_row else 0.0) or 0.0, 3)

    po_not_sap = neg_by_status.get("po_not_sap", 0.0)
    po_awaited = neg_by_status.get("po_awaited", 0.0)
    expected = neg_by_status.get("expected", 0.0)
    return {
        "negotiations": {
            "po_not_sap": po_not_sap,
            "po_awaited": po_awaited,
            "expected": expected,
            "total": round(po_not_sap + po_awaited + expected, 3),
        },
        "business_lost_total": business_lost_total,
    }


def compute(filters):
    enriched = build_enriched()
    if enriched is None:
        return {"error": "No sales data uploaded"}
    po_records = build_po_enriched()

    by = filters.get("by")
    f_team = parse_teams(filters.get("team") or "")
    f_cust = filters.get("customer") or ""
    f_cat = filters.get("category") or ""
    f_country = filters.get("country") or ""
    f_region = filters.get("region") or ""

    all_bys = sorted({e["by"] for e in enriched if e["by"]}, reverse=True)
    if not by:
        by = all_bys[0] if all_bys else None
    pby = prior_by(by) if by else None

    cust_map = get_cust_map()
    mis_map = get_mis_map()

    def match(e):
        if f_team and e["team"] not in f_team: return False
        if f_cust and e["mis_customer"] != f_cust: return False
        if f_cat and e["mis_cat"] != f_cat: return False
        if f_country and e["country"] != f_country: return False
        if f_region and e["region"] != f_region: return False
        return True

    cy = [e for e in enriched if e["by"] == by and match(e)]
    py = [e for e in enriched if e["by"] == pby and match(e)]

    # Auto-detect last available sales month
    auto_up_to = max((e["midx"] for e in cy), default=11)
    up_to_raw = filters.get("up_to_month")
    up_to = auto_up_to if up_to_raw is None or up_to_raw == "" else int(up_to_raw)

    cy_ytd = [e for e in cy if e["midx"] <= up_to]
    py_ytd = [e for e in py if e["midx"] <= up_to]

    # ── PO filter (always for current BY by default; honour same filters) ──
    def match_po(p):
        if f_team and p["team"] not in f_team: return False
        if f_cust and p["mis_customer"] != f_cust: return False
        if f_cat and p["mis_cat"] != f_cat: return False
        if f_country and p["country"] != f_country: return False
        if f_region and p["region"] != f_region: return False
        return True

    po_open = [p for p in po_records if match_po(p)]
    po_total_mn = sum(p["amt_mn"] for p in po_open)

    # ── Targets ──
    targets = (load_dataset("target") or {}).get("records", [])

    def tgt_team(t):
        rec = cust_map.get(t["code"])
        return rec["team"] if rec else ""

    def tgt_region(t):
        rec = cust_map.get(t["code"])
        return (rec.get("region") or "") if rec else ""

    def tgt_mis_cust(t):
        rec = cust_map.get(t["code"])
        if rec and rec.get("mis_customer"):
            return rec["mis_customer"]
        return t.get("name") or ""

    def tgt_mis_cat(t):
        return resolve_mis_category(t.get("third"), t.get("sub_category"),
                                    t.get("category"), mis_map) or "Uncategorised"

    def tgt_match(t):
        if f_team and tgt_team(t) not in f_team: return False
        if f_cust and tgt_mis_cust(t) != f_cust: return False
        if f_cat and tgt_mis_cat(t) != f_cat: return False
        return True

    # YTD & FY target aggregates
    tgt_ytd_total = 0.0
    tgt_fy_total = 0.0
    qtr_target = [0.0, 0.0, 0.0, 0.0]
    for t in targets:
        if not tgt_match(t):
            continue
        for q in range(4):
            qtr_target[q] += t["targets"][q] / 1_000_000.0
        tgt_ytd_total += target_ytd(t["targets"], up_to)
        tgt_fy_total += sum(t["targets"]) / 1_000_000.0

    total_cy = sum(e["sell_mn"] for e in cy_ytd)
    total_py = sum(e["sell_mn"] for e in py_ytd)
    total_py_fy = sum(e["sell_mn"] for e in py)

    # Quarterly and monthly buckets
    qtr_cy = [0.0] * 4
    qtr_py = [0.0] * 4
    mon_cy = [0.0] * 12
    mon_py = [0.0] * 12
    for e in cy:
        mon_cy[e["midx"]] += e["sell_mn"]
        if e["midx"] <= up_to:
            qtr_cy[e["q"]] += e["sell_mn"]
    for e in py:
        mon_py[e["midx"]] += e["sell_mn"]
        qtr_py[e["q"]] += e["sell_mn"]
    qtr_po = [0.0] * 4
    mon_po = [0.0] * 12
    for p in po_open:
        if p.get("q") is not None:
            qtr_po[p["q"]] += p["amt_mn"]
        if p.get("midx") is not None:
            mon_po[p["midx"]] += p["amt_mn"]

    # ── Grouped aggregations ──
    def group(items_cy, items_py, keyfn):
        g = {}
        for e in items_cy:
            k = keyfn(e)
            g.setdefault(k, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
            g[k]["cy"] += e["sell_mn"]
        for e in items_py:
            k = keyfn(e)
            g.setdefault(k, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
            g[k]["py"] += e["sell_mn"]
        return g

    def add_team_targets(g):
        for t in targets:
            if f_cust and tgt_mis_cust(t) != f_cust: continue
            if f_cat and tgt_mis_cat(t) != f_cat: continue
            k = tgt_team(t) or "—"
            g.setdefault(k, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
            g[k]["tgt"] += target_ytd(t["targets"], up_to)

    def add_cust_targets(g):
        for t in targets:
            if f_team and tgt_team(t) not in f_team: continue
            if f_cat and tgt_mis_cat(t) != f_cat: continue
            k = tgt_mis_cust(t) or "—"
            g.setdefault(k, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
            g[k]["tgt"] += target_ytd(t["targets"], up_to)
            g[k].setdefault("tgt_fy", 0.0)
            g[k]["tgt_fy"] += sum(t["targets"]) / 1_000_000.0

    def add_cat_targets(g):
        for t in targets:
            if f_team and tgt_team(t) not in f_team: continue
            if f_cust and tgt_mis_cust(t) != f_cust: continue
            k = tgt_mis_cat(t) or "Uncategorised"
            g.setdefault(k, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
            g[k]["tgt"] += target_ytd(t["targets"], up_to)

    def add_region_targets(g):
        for t in targets:
            if f_team and tgt_team(t) not in f_team: continue
            if f_cust and tgt_mis_cust(t) != f_cust: continue
            if f_cat and tgt_mis_cat(t) != f_cat: continue
            k = tgt_region(t) or "—"
            g.setdefault(k, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
            g[k]["tgt"] += target_ytd(t["targets"], up_to)
            g[k].setdefault("tgt_fy", 0.0)
            g[k]["tgt_fy"] += sum(t["targets"]) / 1_000_000.0

    def add_po(g, keyfn):
        for p in po_open:
            k = keyfn(p)
            g.setdefault(k, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
            g[k]["po"] += p["amt_mn"]

    by_team = group(cy_ytd, py_ytd, lambda e: e["team"])
    add_team_targets(by_team)
    add_po(by_team, lambda p: p["team"])

    by_cust = group(cy_ytd, py_ytd, lambda e: e["mis_customer"])
    add_cust_targets(by_cust)
    add_po(by_cust, lambda p: p["mis_customer"])
    for e in py:
        k = e["mis_customer"]
        by_cust.setdefault(k, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
        by_cust[k].setdefault("py_fy", 0.0)
        by_cust[k]["py_fy"] += e["sell_mn"]

    by_cat = group(cy_ytd, py_ytd, lambda e: e["mis_cat"])
    add_cat_targets(by_cat)
    add_po(by_cat, lambda p: p["mis_cat"])
    # Full-year prior sales per category
    for e in py:
        k = e["mis_cat"]
        by_cat.setdefault(k, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
        by_cat[k].setdefault("py_fy", 0.0)
        by_cat[k]["py_fy"] += e["sell_mn"]
    # Full-year target per category
    for t in targets:
        if f_team and tgt_team(t) not in f_team: continue
        if f_cust and tgt_mis_cust(t) != f_cust: continue
        k = tgt_mis_cat(t) or "Uncategorised"
        by_cat.setdefault(k, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
        by_cat[k].setdefault("tgt_fy", 0.0)
        by_cat[k]["tgt_fy"] += sum(t["targets"]) / 1_000_000.0

    by_country = group(cy_ytd, py_ytd, lambda e: e["country"])
    add_po(by_country, lambda p: p["country"])
    for e in py:
        k = e["country"]
        by_country.setdefault(k, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
        by_country[k].setdefault("py_fy", 0.0)
        by_country[k]["py_fy"] += e["sell_mn"]

    def add_country_targets(g):
        """Allocate each customer's full-year target across destination countries
        in proportion to that customer's actual sales mix (CY+PY) by country."""
        cust_country_mix = {}
        for e in cy + py:
            ck = e["mis_customer"]
            ctry = e["country"]
            cust_country_mix.setdefault(ck, {}).setdefault(ctry, 0.0)
            cust_country_mix[ck][ctry] += e["sell_mn"]
        for t in targets:
            if f_team and tgt_team(t) not in f_team: continue
            if f_cat and tgt_mis_cat(t) != f_cat: continue
            ck = tgt_mis_cust(t) or "—"
            tgt_ytd_val = target_ytd(t["targets"], up_to)
            tgt_fy_val = sum(t["targets"]) / 1_000_000.0
            mix = cust_country_mix.get(ck) or {}
            mix_total = sum(mix.values())
            if mix_total > 0:
                for ctry, amt in mix.items():
                    share = amt / mix_total
                    g.setdefault(ctry, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
                    g[ctry]["tgt"] += tgt_ytd_val * share
                    g[ctry].setdefault("tgt_fy", 0.0)
                    g[ctry]["tgt_fy"] += tgt_fy_val * share
            else:
                g.setdefault("—", {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
                g["—"]["tgt"] += tgt_ytd_val
                g["—"].setdefault("tgt_fy", 0.0)
                g["—"]["tgt_fy"] += tgt_fy_val

    add_country_targets(by_country)

    by_region = group(cy_ytd, py_ytd, lambda e: e["region"])
    add_region_targets(by_region)
    add_po(by_region, lambda p: p["region"])
    for e in py:
        k = e["region"]
        by_region.setdefault(k, {"cy": 0.0, "py": 0.0, "tgt": 0.0, "po": 0.0})
        by_region[k].setdefault("py_fy", 0.0)
        by_region[k]["py_fy"] += e["sell_mn"]

    def serialize(g, limit=None, show_zero=False):
        rows = []
        for k, v in g.items():
            if not show_zero and v["cy"] == 0 and v["py"] == 0 and v["tgt"] == 0 and v["po"] == 0 \
                    and v.get("tgt_fy", 0.0) == 0 and v.get("py_fy", 0.0) == 0:
                continue
            ach = (v["cy"] / v["tgt"] * 100) if v["tgt"] > 0 else None
            yoy = ((v["cy"] - v["py"]) / v["py"] * 100) if v["py"] > 0 else None
            rows.append({
                "name": k or "—",
                "cy": round(v["cy"], 3),
                "py": round(v["py"], 3),
                "tgt": round(v["tgt"], 3),
                "po": round(v["po"], 3),
                "ach": round(ach, 1) if ach is not None else None,
                "yoy": round(yoy, 1) if yoy is not None else None,
                "tgt_fy": round(v.get("tgt_fy", 0.0), 3),
                "py_fy": round(v.get("py_fy", 0.0), 3),
            })
        rows.sort(key=lambda r: r["cy"] + r["po"], reverse=True)
        return rows[:limit] if limit else rows

    all_teams = sort_teams({v["team"] for v in cust_map.values()
                            if v.get("team") and v["team"] != "—"})

    # Cascading dropdowns
    team_scope = [e for e in enriched if e["by"] == by and (not f_team or e["team"] in f_team)]
    filtered_custs = sorted({e["mis_customer"] for e in team_scope if e["mis_customer"] != "—"})
    filtered_countries = sorted({e["country"] for e in team_scope if e["country"] != "—"})

    # Categories that the selected team has a target for
    tgt_cats_set = set()
    for t in targets:
        if f_team and tgt_team(t) not in f_team:
            continue
        mc = tgt_mis_cat(t)
        if mc:
            tgt_cats_set.add(mc)
    tgt_cats = sorted(tgt_cats_set)

    all_cats = sorted({e["mis_cat"] for e in enriched if e["mis_cat"]})
    all_regions = sorted({e["region"] for e in enriched if e["region"] != "—"})

    # No-sale lists: had PY sales but zero CY sales (same YTD window)
    cy_cust_set = {e["mis_customer"] for e in cy_ytd if e["mis_customer"] != "—"}
    py_cust_sales = {}
    for e in py_ytd:
        if e["mis_customer"] != "—":
            py_cust_sales[e["mis_customer"]] = py_cust_sales.get(e["mis_customer"], 0.0) + e["sell_mn"]
    no_sale_customers = sorted(
        [{"name": k, "py": round(v, 3)} for k, v in py_cust_sales.items() if k not in cy_cust_set],
        key=lambda r: -r["py"]
    )

    cy_cat_set = {e["mis_cat"] for e in cy_ytd if e["mis_cat"] != "Uncategorised"}
    py_cat_sales = {}
    for e in py_ytd:
        if e["mis_cat"] and e["mis_cat"] != "Uncategorised":
            py_cat_sales[e["mis_cat"]] = py_cat_sales.get(e["mis_cat"], 0.0) + e["sell_mn"]
    no_sale_categories = sorted(
        [{"name": k, "py": round(v, 3)} for k, v in py_cat_sales.items() if k not in cy_cat_set],
        key=lambda r: -r["py"]
    )

    yoy_total = ((total_cy - total_py) / total_py * 100) if total_py > 0 else None
    ach_total = (total_cy / tgt_ytd_total * 100) if tgt_ytd_total > 0 else None
    ach_plus_po = ((total_cy + po_total_mn) / tgt_fy_total * 100) if tgt_fy_total > 0 else None

    return {
        "by": by, "prior_by": pby, "up_to_month": up_to,
        "month_label": MONTH_LABELS[up_to],
        "last_sales_month": auto_up_to,
        "last_sales_month_label": MONTH_LABELS[auto_up_to],
        "all_bys": all_bys,
        "kpis": {
            "sales_cy": round(total_cy, 3),
            "sales_py": round(total_py, 3),
            "sales_py_fy": round(total_py_fy, 3),
            "yoy": round(yoy_total, 1) if yoy_total is not None else None,
            "target_ytd": round(tgt_ytd_total, 3),
            "target_fy": round(tgt_fy_total, 3),
            "ach": round(ach_total, 1) if ach_total is not None else None,
            "gap": round(tgt_fy_total - total_cy, 3),
            "months_left": 11 - up_to,
            "po_open": round(po_total_mn, 3),
            "sales_plus_po": round(total_cy + po_total_mn, 3),
            "gap_after_po": round(tgt_fy_total - total_cy - po_total_mn, 3),
            "ach_with_po": round(ach_plus_po, 1) if ach_plus_po is not None else None,
        },
        "quarterly": {
            "labels": QUARTER_LABELS,
            "cy": [round(x, 3) for x in qtr_cy],
            "py": [round(x, 3) for x in qtr_py],
            "target": [round(x, 3) for x in qtr_target],
            "po": [round(x, 3) for x in qtr_po],
        },
        "monthly": {
            "labels": MONTH_LABELS,
            "cy": [round(x, 3) if i <= up_to else None for i, x in enumerate(mon_cy)],
            "py": [round(x, 3) for x in mon_py],
            "po": [round(x, 3) if x > 0 else None for x in mon_po],
        },
        "no_sale_customers": no_sale_customers,
        "no_sale_categories": no_sale_categories,
        "by_team": serialize(by_team),
        "by_customer": serialize(by_cust),
        "by_category": serialize(by_cat),
        "by_country": serialize(by_country, 40),
        "by_region": serialize(by_region),
        "notes_summary": notes_overall_summary(by, f_team),
        "options": {
            "teams": all_teams,
            "customers": filtered_custs,
            "categories": all_cats,
            "countries": filtered_countries,
            "regions": all_regions,
            "tgt_categories": tgt_cats,
        },
    }


# ──────────────────────────────────────────────────────────────────────
#  Auth helpers
# ──────────────────────────────────────────────────────────────────────
def get_current_user():
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
    else:
        # Allow token via query param for browser-initiated downloads
        token = request.args.get("token", "")
    if not token:
        return None
    with closing(get_db()) as db:
        row = db.execute("SELECT * FROM tokens WHERE token=?", (token,)).fetchone()
    return dict(row) if row else None


def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not get_current_user():
            return jsonify({"error": "unauthorized"}), 401
        return f(*a, **kw)
    return dec


def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        u = get_current_user()
        if not u:
            return jsonify({"error": "unauthorized"}), 401
        if u["role"] != "admin":
            return jsonify({"error": "forbidden"}), 403
        return f(*a, **kw)
    return dec


def analyst_required(f):
    """Allows admin or business-analyst roles (used for Meeting Minutes edits)."""
    @wraps(f)
    def dec(*a, **kw):
        u = get_current_user()
        if not u:
            return jsonify({"error": "unauthorized"}), 401
        if u["role"] not in ("admin", "analyst"):
            return jsonify({"error": "forbidden"}), 403
        return f(*a, **kw)
    return dec


# ──────────────────────────────────────────────────────────────────────
#  Auth routes
# ──────────────────────────────────────────────────────────────────────
@app.route("/api/auth/me")
def auth_me():
    u = get_current_user()
    if not u:
        return jsonify({"authenticated": False})
    return jsonify({"authenticated": True, "username": u["username"],
                    "role": u["role"], "team": u["team"] or ""})


@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json(force=True, silent=True) or {}
    username = str(data.get("username", "")).strip()
    password = str(data.get("password", ""))
    with closing(get_db()) as db:
        row = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not row or not check_password_hash(row["password"], password):
        return jsonify({"error": "Invalid username or password"}), 401
    token = secrets.token_urlsafe(32)
    with closing(get_db()) as db:
        db.execute("INSERT INTO tokens (token,user_id,username,role,team,created_at) VALUES(?,?,?,?,?,?)",
                   (token, row["id"], row["username"], row["role"], row["team"] or "",
                    dt.datetime.now().isoformat(timespec="seconds")))
        db.commit()
    return jsonify({"ok": True, "token": token, "username": row["username"],
                    "role": row["role"], "team": row["team"] or ""})


@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        with closing(get_db()) as db:
            db.execute("DELETE FROM tokens WHERE token=?", (auth[7:],))
            db.commit()
    return jsonify({"ok": True})


@app.route("/api/auth/change-password", methods=["POST"])
def auth_change_password():
    data = request.get_json(force=True, silent=True) or {}
    username = str(data.get("username", "")).strip()
    current_password = str(data.get("current_password", ""))
    new_password = str(data.get("new_password", ""))
    if not username or not current_password or not new_password:
        return jsonify({"error": "Username, current password and new password are required"}), 400
    if len(new_password) < 4:
        return jsonify({"error": "New password must be at least 4 characters"}), 400
    with closing(get_db()) as db:
        row = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        if not row or not check_password_hash(row["password"], current_password):
            return jsonify({"error": "Invalid username or current password"}), 401
        db.execute("UPDATE users SET password=? WHERE id=?",
                   (generate_password_hash(new_password), row["id"]))
        db.execute("DELETE FROM tokens WHERE user_id=?", (row["id"],))
        db.commit()
    return jsonify({"ok": True})


# ──────────────────────────────────────────────────────────────────────
#  Admin user management
# ──────────────────────────────────────────────────────────────────────
@app.route("/api/admin/teams", methods=["GET"])
@admin_required
def admin_teams():
    """Return unique team names from the customer master (or team-market config as fallback)."""
    teams = set()
    with closing(get_db()) as db:
        row = db.execute("SELECT payload FROM datasets WHERE key='cust_team_master'").fetchone()
        if row:
            try:
                records = json.loads(row["payload"])
                for r in records:
                    t = str(r.get("team") or "").strip()
                    if t and t != "—":
                        teams.add(t)
            except Exception:
                pass
        # Fallback: use mi_team_markets team names if customer master not uploaded yet
        if not teams:
            rows = db.execute(
                "SELECT DISTINCT team_name FROM mi_team_markets WHERE is_active=1 ORDER BY team_name"
            ).fetchall()
            for r in rows:
                teams.add(r["team_name"])
    return jsonify(sort_teams(teams))


@app.route("/api/admin/users", methods=["GET"])
@admin_required
def admin_list():
    with closing(get_db()) as db:
        rows = db.execute("SELECT id,username,role,team,created_at FROM users ORDER BY id").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/users", methods=["POST"])
@admin_required
def admin_create():
    data = request.get_json(force=True, silent=True) or {}
    u = str(data.get("username", "")).strip()
    p = str(data.get("password", "")).strip()
    role = data.get("role", "team")
    team = str(data.get("team", "")).strip()
    if not u or not p:
        return jsonify({"error": "username and password required"}), 400
    if role not in ("admin", "team"):
        return jsonify({"error": "invalid role"}), 400
    try:
        with closing(get_db()) as db:
            db.execute("INSERT INTO users (username,password,role,team,created_at) VALUES(?,?,?,?,?)",
                       (u, generate_password_hash(p), role, team,
                        dt.datetime.now().isoformat(timespec="seconds")))
            db.commit()
            uid = db.execute("SELECT id FROM users WHERE username=?", (u,)).fetchone()["id"]
        return jsonify({"ok": True, "id": uid})
    except sqlite3.IntegrityError:
        return jsonify({"error": "Username already exists"}), 409


@app.route("/api/admin/users/<int:uid>", methods=["PUT"])
@admin_required
def admin_update(uid):
    data = request.get_json(force=True, silent=True) or {}
    role = data.get("role", "team")
    team = str(data.get("team", "")).strip()
    if role not in ("admin", "team"):
        return jsonify({"error": "invalid role"}), 400
    with closing(get_db()) as db:
        if data.get("password"):
            db.execute("UPDATE users SET role=?,team=?,password=? WHERE id=?",
                       (role, team, generate_password_hash(data["password"]), uid))
        else:
            db.execute("UPDATE users SET role=?,team=? WHERE id=?", (role, team, uid))
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/users/<int:uid>", methods=["DELETE"])
@admin_required
def admin_delete(uid):
    u = get_current_user()
    if u and u["user_id"] == uid:
        return jsonify({"error": "Cannot delete your own account"}), 400
    with closing(get_db()) as db:
        db.execute("DELETE FROM users WHERE id=?", (uid,))
        db.execute("DELETE FROM tab_access WHERE user_id=?", (uid,))
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/tab-access", methods=["GET"])
@admin_required
def admin_tab_access_get():
    """Return current tab access: {pm: [user_ids], mn: [user_ids], mi: [user_ids]}"""
    with closing(get_db()) as db:
        rows = db.execute("SELECT tab_name, user_id FROM tab_access").fetchall()
    result = {"pm": [], "mn": [], "mi": []}
    for r in rows:
        if r["tab_name"] in result:
            result[r["tab_name"]].append(r["user_id"])
    return jsonify(result)


@app.route("/api/admin/tab-access", methods=["POST"])
@admin_required
def admin_tab_access_set():
    """Replace access list for a tab. Body: {tab: 'pm'|'mn'|'mi', user_ids: [int,...]}"""
    data = request.get_json(force=True, silent=True) or {}
    tab = str(data.get("tab", "")).strip()
    if tab not in ("pm", "mn", "mi"):
        return jsonify({"error": "Invalid tab"}), 400
    user_ids = [int(x) for x in (data.get("user_ids") or []) if str(x).isdigit() or isinstance(x, int)]
    with closing(get_db()) as db:
        db.execute("DELETE FROM tab_access WHERE tab_name=?", (tab,))
        for uid in user_ids:
            db.execute("INSERT OR IGNORE INTO tab_access (tab_name, user_id) VALUES (?,?)", (tab, uid))
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/tab-access", methods=["GET"])
@login_required
def user_tab_access():
    """Return which tabs the current (non-admin) user can see."""
    u = get_current_user()
    if not u:
        return jsonify({"error": "unauthorized"}), 401
    uid = u["user_id"]
    with closing(get_db()) as db:
        rows = db.execute("SELECT tab_name FROM tab_access WHERE user_id=?", (uid,)).fetchall()
    granted = {r["tab_name"] for r in rows}
    return jsonify({"pm": "pm" in granted, "mn": "mn" in granted, "mi": "mi" in granted})


@app.route("/api/admin/reconciliation", methods=["GET"])
@admin_required
def admin_reconciliation():
    """Return data-mismatch summary for admin panel:
    - Unmatched customer codes (Sales, PO, Target)
    - Unmatched categories (Sales, PO, Target)
    """
    cust_map = get_cust_map()
    mis_map  = get_mis_map()

    # ── helpers ──────────────────────────────────────────────────────────
    def _cat_key(third, sub, cat):
        return " / ".join(filter(None, [third or "", sub or "", cat or ""])) or "(blank)"

    # ── Sales ────────────────────────────────────────────────────────────
    sales_ds = load_dataset("sales")
    unmatched_cust_sales  = {}   # code → {sales_mn, rows, sample_name}
    unmatched_cat_sales   = {}   # cat_key → {sales_mn, rows, codes}

    if sales_ds:
        for r in sales_ds["records"]:
            code = r.get("code", "")
            if code and code not in cust_map:
                if code not in unmatched_cust_sales:
                    unmatched_cust_sales[code] = {"sales_mn": 0.0, "rows": 0, "source": "Sales"}
                unmatched_cust_sales[code]["sales_mn"] += r.get("sell_mn", 0.0)
                unmatched_cust_sales[code]["rows"] += 1
            cat_k = _cat_key(r.get("third"), r.get("sub_category"), r.get("category"))
            if not resolve_mis_category(r.get("third"), r.get("sub_category"), r.get("category"), mis_map):
                if cat_k not in unmatched_cat_sales:
                    unmatched_cat_sales[cat_k] = {"sales_mn": 0.0, "rows": 0, "codes": set()}
                unmatched_cat_sales[cat_k]["sales_mn"] += r.get("sell_mn", 0.0)
                unmatched_cat_sales[cat_k]["rows"] += 1
                unmatched_cat_sales[cat_k]["codes"].add(code or "—")

    # ── PO ───────────────────────────────────────────────────────────────
    po_ds = load_dataset("po")
    unmatched_cust_po  = {}
    unmatched_cat_po   = {}

    if po_ds:
        for r in po_ds["records"]:
            code = r.get("code", "")
            if code and code not in cust_map:
                if code not in unmatched_cust_po:
                    unmatched_cust_po[code] = {"amt_mn": 0.0, "rows": 0, "source": "PO"}
                unmatched_cust_po[code]["amt_mn"] += r.get("amt_mn", 0.0)
                unmatched_cust_po[code]["rows"] += 1
            cat_k = _cat_key(r.get("third"), r.get("sub_category"), r.get("category"))
            if not resolve_mis_category(r.get("third"), r.get("sub_category"), r.get("category"), mis_map):
                if cat_k not in unmatched_cat_po:
                    unmatched_cat_po[cat_k] = {"amt_mn": 0.0, "rows": 0, "codes": set()}
                unmatched_cat_po[cat_k]["amt_mn"] += r.get("amt_mn", 0.0)
                unmatched_cat_po[cat_k]["rows"] += 1
                unmatched_cat_po[cat_k]["codes"].add(code or "—")

    # ── Target ───────────────────────────────────────────────────────────
    tgt_records = (load_dataset("target") or {}).get("records", [])
    unmatched_cust_tgt  = {}
    unmatched_cat_tgt   = {}

    for r in tgt_records:
        code = r.get("code", "")
        tgt_total = sum(r.get("targets", [0, 0, 0, 0]))
        if code and code not in cust_map:
            if code not in unmatched_cust_tgt:
                unmatched_cust_tgt[code] = {
                    "tgt_mn": 0.0, "rows": 0,
                    "name": r.get("name") or "", "source": "Target"
                }
            unmatched_cust_tgt[code]["tgt_mn"] += tgt_total / 1_000_000.0
            unmatched_cust_tgt[code]["rows"] += 1
        cat_k = _cat_key(r.get("third"), r.get("sub_category"), r.get("category"))
        if not resolve_mis_category(r.get("third"), r.get("sub_category"), r.get("category"), mis_map):
            if cat_k not in unmatched_cat_tgt:
                unmatched_cat_tgt[cat_k] = {"tgt_mn": 0.0, "rows": 0, "codes": set()}
            unmatched_cat_tgt[cat_k]["tgt_mn"] += tgt_total / 1_000_000.0
            unmatched_cat_tgt[cat_k]["rows"] += 1
            unmatched_cat_tgt[cat_k]["codes"].add(code or "—")

    # ── Serialise ────────────────────────────────────────────────────────
    def _cust_rows(sales_d, po_d, tgt_d):
        codes = sorted(set(list(sales_d) + list(po_d) + list(tgt_d)))
        out = []
        for code in codes:
            out.append({
                "code":     code,
                "sales_mn": round(sales_d.get(code, {}).get("sales_mn", 0.0), 3),
                "po_mn":    round(po_d.get(code, {}).get("amt_mn", 0.0), 3),
                "tgt_mn":   round(tgt_d.get(code, {}).get("tgt_mn", 0.0), 3),
                "name":     tgt_d.get(code, {}).get("name", ""),
                "rows":     (sales_d.get(code, {}).get("rows", 0)
                             + po_d.get(code, {}).get("rows", 0)
                             + tgt_d.get(code, {}).get("rows", 0)),
            })
        out.sort(key=lambda x: -(x["sales_mn"] + x["po_mn"] + x["tgt_mn"]))
        return out

    def _cat_rows(sales_d, po_d, tgt_d):
        keys = sorted(set(list(sales_d) + list(po_d) + list(tgt_d)))
        out = []
        for k in keys:
            out.append({
                "category":  k,
                "sales_mn":  round(sales_d.get(k, {}).get("sales_mn", 0.0), 3),
                "po_mn":     round(po_d.get(k, {}).get("amt_mn", 0.0), 3),
                "tgt_mn":    round(tgt_d.get(k, {}).get("tgt_mn", 0.0), 3),
                "rows":      (sales_d.get(k, {}).get("rows", 0)
                              + po_d.get(k, {}).get("rows", 0)
                              + tgt_d.get(k, {}).get("rows", 0)),
                "codes":     sorted((sales_d.get(k, {}).get("codes", set())
                                     | po_d.get(k, {}).get("codes", set())
                                     | tgt_d.get(k, {}).get("codes", set()))),
            })
        out.sort(key=lambda x: -(x["sales_mn"] + x["po_mn"] + x["tgt_mn"]))
        return out

    return jsonify({
        "unmatched_customers": _cust_rows(unmatched_cust_sales, unmatched_cust_po, unmatched_cust_tgt),
        "unmatched_categories": _cat_rows(unmatched_cat_sales, unmatched_cat_po, unmatched_cat_tgt),
        "has_sales":  sales_ds is not None,
        "has_po":     po_ds is not None,
        "has_target": len(tgt_records) > 0,
    })


@app.route("/api/admin/uncategorised/download", methods=["GET"])
@admin_required
def admin_uncategorised_download():
    """Return an Excel workbook of sales rows whose MIS Category is unresolved."""
    sales_ds = load_dataset("sales")
    if not sales_ds:
        return jsonify({"error": "No sales data uploaded"}), 400

    mis_map = get_mis_map()
    cust_map = get_cust_map()

    groups = {}  # (third, sub, cat) → {sales_mn, count, customers}
    for r in sales_ds["records"]:
        third = r.get("third", "")
        sub   = r.get("sub_category", "")
        cat   = r.get("category", "")
        if resolve_mis_category(third, sub, cat, mis_map):
            continue
        cm = cust_map.get(r.get("code", ""), {})
        cust_name = cm.get("mis_customer") or cm.get("name") or r.get("code", "") or "—"
        key = (third or "(blank)", sub or "(blank)", cat or "(blank)")
        if key not in groups:
            groups[key] = {"sales_mn": 0.0, "count": 0, "customers": set()}
        groups[key]["sales_mn"] += r.get("sell_mn", 0.0)
        groups[key]["count"] += 1
        groups[key]["customers"].add(cust_name)

    rows = []
    for (third, sub, cat), v in groups.items():
        rows.append({
            "Product Third Category": third,
            "Product Sub Category":   sub,
            "Product Category":       cat,
            "Total Sales (USD Mn)":   round(v["sales_mn"], 3),
            "Invoice Lines":          v["count"],
            "Customers":              ", ".join(sorted(v["customers"])),
        })
    rows.sort(key=lambda x: x["Total Sales (USD Mn)"], reverse=True)

    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["Product Third Category", "Product Sub Category",
                 "Product Category", "Total Sales (USD Mn)", "Invoice Lines", "Customers"])

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Uncategorised")
        ws = writer.sheets["Uncategorised"]
        for col_cells in ws.columns:
            width = max(len(str(c.value or "")) for c in col_cells) + 4
            ws.column_dimensions[col_cells[0].column_letter].width = min(width, 60)
    bio.seek(0)

    from datetime import date as _date
    fname = f"uncategorised_{_date.today().isoformat()}.xlsx"
    return send_file(bio,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=fname)


@app.route("/api/admin/zecost06/qc-download", methods=["GET"])
@admin_required
def admin_zecost06_qc_download():
    """Download the parsed zecost06 records as Excel for QC verification."""
    ds = load_dataset("zecost06")
    if not ds:
        return jsonify({"error": "zecost06 not uploaded"}), 400

    cust_map = get_cust_map()
    rows = []
    for r in ds["records"]:
        cm = cust_map.get(r.get("customer_code", ""), {})
        rows.append({
            "Customer Code":        r.get("customer_code", ""),
            "Customer Name":        cm.get("name") or cm.get("mis_customer") or "",
            "Export ID":            r.get("export_id", ""),
            "Material":             r.get("material", ""),
            "PO Qty":               r.get("po_qty", 0),
            "PO Cost (USD)":        r.get("po_cost_usd", 0),
            "PO Sale Value (USD)":  r.get("po_sale_usd", 0),
            "Quot Qty":             r.get("quot_qty", 0),
            "Quot Cost (USD)":      r.get("quot_cost_usd", 0),
            "Quot Sale Value (USD)":r.get("quot_sale_usd", 0),
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame()
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Parsed Cost Data")
        ws = writer.sheets["Parsed Cost Data"]
        for col_cells in ws.columns:
            width = max(len(str(c.value or "")) for c in col_cells) + 4
            ws.column_dimensions[col_cells[0].column_letter].width = min(width, 40)
    bio.seek(0)

    from datetime import date as _date
    fname = f"zecost06_parsed_{_date.today().isoformat()}.xlsx"
    return send_file(bio,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=fname)


# ──────────────────────────────────────────────────────────────────────
#  Team notes (negotiations & challenges)
# ──────────────────────────────────────────────────────────────────────
def _current_month_key():
    return dt.date.today().strftime("%Y-%m")


@app.route("/api/notes", methods=["GET"])
@login_required
def notes_list():
    u = get_current_user()
    by = request.args.get("by", "")
    ntype = request.args.get("type", "")
    team = resolve_team_filter(u, request.args.get("team", ""))
    month = request.args.get("month", "")  # specific month_key, else current month
    if not month:
        month = _current_month_key()
    with closing(get_db()) as db:
        q = "SELECT * FROM team_note_items WHERE 1=1"
        params = []
        if by:
            q += " AND by=?"; params.append(by)
        if ntype:
            q += " AND particulars=?"; params.append(ntype)
        if ntype in ("negotiation", "challenge"):
            q += " AND month_key=?"; params.append(month)
        if team:
            team_set = parse_teams(team)
            placeholders = ",".join("?" * len(team_set))
            q += f" AND team IN ({placeholders})"
            params.extend(sorted(team_set))
        rows = db.execute(q + " ORDER BY id", params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/notes/archive", methods=["GET"])
@login_required
def notes_archive():
    """Return past months' archived snapshots for negotiations/challenges,
    grouped by month_key (most recent first), excluding the current month."""
    u = get_current_user()
    by = request.args.get("by", "")
    ntype = request.args.get("type", "")
    team = resolve_team_filter(u, request.args.get("team", ""))
    if ntype not in ("negotiation", "challenge"):
        return jsonify({"error": "type must be negotiation or challenge"}), 400
    cur_month = _current_month_key()
    with closing(get_db()) as db:
        q = "SELECT * FROM team_note_items WHERE particulars=? AND month_key!='' AND month_key!=?"
        params = [ntype, cur_month]
        if by:
            q += " AND by=?"; params.append(by)
        if team:
            team_set = parse_teams(team)
            placeholders = ",".join("?" * len(team_set))
            q += f" AND team IN ({placeholders})"
            params.extend(sorted(team_set))
        rows = db.execute(q + " ORDER BY month_key DESC, id", params).fetchall()
    months = {}
    for r in rows:
        d = dict(r)
        months.setdefault(d["month_key"], []).append(d)
    result = [{"month": m, "items": items} for m, items in
               sorted(months.items(), key=lambda x: x[0], reverse=True)]
    return jsonify(result)


@app.route("/api/notes", methods=["POST"])
@login_required
def notes_create():
    u = get_current_user()
    data = request.get_json(force=True, silent=True) or {}
    by = str(data.get("by", "")).strip()
    team = str(data.get("team", "")).strip()
    customer = str(data.get("customer", "")).strip()
    particulars = str(data.get("particulars", "")).strip()
    try:
        value_usd = float(data.get("value_usd") or 0)
    except (ValueError, TypeError):
        value_usd = 0.0
    notes = str(data.get("notes", "")).strip()
    program = str(data.get("program", "")).strip()
    period = str(data.get("period", "")).strip()
    status = str(data.get("status", "po_not_sap")).strip() if particulars == "negotiation" else ""
    if not by or not team or particulars not in ("negotiation", "challenge", "business_lost"):
        return jsonify({"error": "by, team, and valid particulars required"}), 400
    if u and u["role"] == "team" and u["team"] and team not in parse_teams(u["team"]):
        return jsonify({"error": "forbidden"}), 403
    with closing(get_db()) as db:
        db.execute(
            "INSERT INTO team_note_items (by,team,customer,particulars,value_usd,notes,status,program,period,month_key,updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (by, team, customer, particulars, value_usd, notes, status, program, period,
             _current_month_key(), dt.datetime.now().isoformat(timespec="seconds"))
        )
        db.commit()
        nid = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    return jsonify({"ok": True, "id": nid}), 201


@app.route("/api/notes/<int:nid>", methods=["PUT"])
@login_required
def notes_update(nid):
    u = get_current_user()
    with closing(get_db()) as db:
        existing = db.execute("SELECT * FROM team_note_items WHERE id=?", (nid,)).fetchone()
    if not existing:
        return jsonify({"error": "not found"}), 404
    existing = dict(existing)
    if u and u["role"] == "team" and u["team"] and existing["team"] not in parse_teams(u["team"]):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json(force=True, silent=True) or {}
    customer = str(data.get("customer", existing["customer"])).strip()
    try:
        value_usd = float(data.get("value_usd") or 0)
    except (ValueError, TypeError):
        value_usd = existing["value_usd"]
    notes = str(data.get("notes", existing["notes"] or "")).strip()
    status = str(data.get("status", existing.get("status") or "po_not_sap")).strip()
    program = str(data.get("program", existing.get("program") or "")).strip()
    period = str(data.get("period", existing.get("period") or "")).strip()
    cur_month = _current_month_key()
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        if existing["particulars"] == "negotiation" and existing.get("month_key") and existing["month_key"] != cur_month:
            # Month has rolled over since this row was last edited — leave the
            # old row as an archived snapshot and create a fresh one for this month.
            db.execute(
                "INSERT INTO team_note_items (by,team,customer,particulars,value_usd,notes,status,program,period,month_key,updated_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (existing["by"], existing["team"], customer, existing["particulars"], value_usd, notes,
                 status, program, period, cur_month, now)
            )
            db.commit()
            nid = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
            return jsonify({"ok": True, "id": nid, "rolled_over": True}), 201
        db.execute(
            "UPDATE team_note_items SET customer=?,value_usd=?,notes=?,status=?,program=?,period=?,updated_at=? WHERE id=?",
            (customer, value_usd, notes, status, program, period, now, nid)
        )
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/notes/<int:nid>", methods=["DELETE"])
@login_required
def notes_delete(nid):
    u = get_current_user()
    with closing(get_db()) as db:
        existing = db.execute("SELECT * FROM team_note_items WHERE id=?", (nid,)).fetchone()
    if not existing:
        return jsonify({"error": "not found"}), 404
    existing = dict(existing)
    if u and u["role"] == "team" and u["team"] and existing["team"] not in parse_teams(u["team"]):
        return jsonify({"error": "forbidden"}), 403
    with closing(get_db()) as db:
        db.execute("DELETE FROM team_note_items WHERE id=?", (nid,))
        db.commit()
    return jsonify({"ok": True})


# ──────────────────────────────────────────────────────────────────────
#  Meeting minutes (monthly sales review)
# ──────────────────────────────────────────────────────────────────────
def _minute_with_items(db, minute):
    d = dict(minute)
    items = db.execute(
        "SELECT * FROM meeting_action_items WHERE minute_id=? ORDER BY id", (d["id"],)
    ).fetchall()
    d["action_items"] = [dict(r) for r in items]
    return d


@app.route("/api/meeting-minutes", methods=["GET"])
@login_required
def meeting_minutes_list():
    by = request.args.get("by", "")
    with closing(get_db()) as db:
        q = "SELECT * FROM meeting_minutes WHERE 1=1"
        params = []
        if by:
            q += " AND by=?"; params.append(by)
        rows = db.execute(q + " ORDER BY meeting_date DESC, id DESC", params).fetchall()
        result = [_minute_with_items(db, r) for r in rows]
    return jsonify(result)


@app.route("/api/meeting-minutes", methods=["POST"])
@analyst_required
def meeting_minutes_create():
    u = get_current_user()
    data = request.get_json(force=True, silent=True) or {}
    by = str(data.get("by", "")).strip()
    meeting_date = str(data.get("meeting_date", "")).strip()
    title = str(data.get("title", "")).strip()
    attendees = str(data.get("attendees", "")).strip()
    circulated_to = str(data.get("circulated_to", "")).strip()
    notes = str(data.get("notes", "")).strip()
    action_items = data.get("action_items") or []
    if not by or not meeting_date:
        return jsonify({"error": "by and meeting_date required"}), 400
    month_key = meeting_date[:7]
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        db.execute(
            "INSERT INTO meeting_minutes (by,month_key,meeting_date,title,attendees,circulated_to,notes,created_by,created_at,updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (by, month_key, meeting_date, title, attendees, circulated_to, notes,
             u["username"] if u else "", now, now)
        )
        db.commit()
        mid = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        for ai in action_items:
            description = str(ai.get("description", "")).strip()
            if not description:
                continue
            responsible = str(ai.get("responsible", "")).strip()
            deadline = str(ai.get("deadline", "")).strip()
            ai_status = str(ai.get("status", "open")).strip() or "open"
            db.execute(
                "INSERT INTO meeting_action_items (minute_id,description,responsible,deadline,status,created_at,updated_at) "
                "VALUES (?,?,?,?,?,?,?)",
                (mid, description, responsible, deadline, ai_status, now, now)
            )
        db.commit()
    return jsonify({"ok": True, "id": mid}), 201


@app.route("/api/meeting-minutes/<int:mid>", methods=["PUT"])
@analyst_required
def meeting_minutes_update(mid):
    with closing(get_db()) as db:
        existing = db.execute("SELECT * FROM meeting_minutes WHERE id=?", (mid,)).fetchone()
    if not existing:
        return jsonify({"error": "not found"}), 404
    existing = dict(existing)
    data = request.get_json(force=True, silent=True) or {}
    meeting_date = str(data.get("meeting_date", existing["meeting_date"])).strip()
    title = str(data.get("title", existing["title"] or "")).strip()
    attendees = str(data.get("attendees", existing["attendees"] or "")).strip()
    circulated_to = str(data.get("circulated_to", existing["circulated_to"] or "")).strip()
    notes = str(data.get("notes", existing["notes"] or "")).strip()
    month_key = meeting_date[:7]
    now = dt.datetime.now().isoformat(timespec="seconds")
    action_items = data.get("action_items")
    with closing(get_db()) as db:
        db.execute(
            "UPDATE meeting_minutes SET meeting_date=?,month_key=?,title=?,attendees=?,circulated_to=?,notes=?,updated_at=? WHERE id=?",
            (meeting_date, month_key, title, attendees, circulated_to, notes, now, mid)
        )
        if action_items is not None:
            # Replace the full action-item list with the one submitted by the form.
            db.execute("DELETE FROM meeting_action_items WHERE minute_id=?", (mid,))
            for ai in action_items:
                description = str(ai.get("description", "")).strip()
                if not description:
                    continue
                responsible = str(ai.get("responsible", "")).strip()
                deadline = str(ai.get("deadline", "")).strip()
                ai_status = str(ai.get("status", "open")).strip() or "open"
                db.execute(
                    "INSERT INTO meeting_action_items (minute_id,description,responsible,deadline,status,created_at,updated_at) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (mid, description, responsible, deadline, ai_status, now, now)
                )
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/meeting-minutes/<int:mid>", methods=["DELETE"])
@analyst_required
def meeting_minutes_delete(mid):
    with closing(get_db()) as db:
        existing = db.execute("SELECT * FROM meeting_minutes WHERE id=?", (mid,)).fetchone()
        if not existing:
            return jsonify({"error": "not found"}), 404
        db.execute("DELETE FROM meeting_action_items WHERE minute_id=?", (mid,))
        db.execute("DELETE FROM meeting_minutes WHERE id=?", (mid,))
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/meeting-minutes/<int:mid>/action-items", methods=["POST"])
@analyst_required
def meeting_action_item_create(mid):
    with closing(get_db()) as db:
        existing = db.execute("SELECT id FROM meeting_minutes WHERE id=?", (mid,)).fetchone()
    if not existing:
        return jsonify({"error": "not found"}), 404
    data = request.get_json(force=True, silent=True) or {}
    description = str(data.get("description", "")).strip()
    if not description:
        return jsonify({"error": "description required"}), 400
    responsible = str(data.get("responsible", "")).strip()
    deadline = str(data.get("deadline", "")).strip()
    status = str(data.get("status", "open")).strip() or "open"
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        db.execute(
            "INSERT INTO meeting_action_items (minute_id,description,responsible,deadline,status,created_at,updated_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (mid, description, responsible, deadline, status, now, now)
        )
        db.commit()
        aid = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    return jsonify({"ok": True, "id": aid}), 201


@app.route("/api/meeting-action-items/<int:aid>", methods=["PUT"])
@login_required
def meeting_action_item_update(aid):
    """Description/responsible/deadline edits require admin/analyst; any
    logged-in user may update only the status (to mark progress)."""
    u = get_current_user()
    with closing(get_db()) as db:
        existing = db.execute("SELECT * FROM meeting_action_items WHERE id=?", (aid,)).fetchone()
    if not existing:
        return jsonify({"error": "not found"}), 404
    existing = dict(existing)
    data = request.get_json(force=True, silent=True) or {}
    is_analyst = u and u["role"] in ("admin", "analyst")
    if is_analyst:
        description = str(data.get("description", existing["description"])).strip()
        responsible = str(data.get("responsible", existing["responsible"] or "")).strip()
        deadline = str(data.get("deadline", existing["deadline"] or "")).strip()
    else:
        description = existing["description"]
        responsible = existing["responsible"]
        deadline = existing["deadline"]
    status = str(data.get("status", existing["status"] or "open")).strip() or "open"
    if status not in ("open", "in_progress", "done"):
        status = existing["status"] or "open"
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        db.execute(
            "UPDATE meeting_action_items SET description=?,responsible=?,deadline=?,status=?,updated_at=? WHERE id=?",
            (description, responsible, deadline, status, now, aid)
        )
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/meeting-action-items/<int:aid>", methods=["DELETE"])
@analyst_required
def meeting_action_item_delete(aid):
    with closing(get_db()) as db:
        existing = db.execute("SELECT id FROM meeting_action_items WHERE id=?", (aid,)).fetchone()
        if not existing:
            return jsonify({"error": "not found"}), 404
        db.execute("DELETE FROM meeting_action_items WHERE id=?", (aid,))
        db.commit()
    return jsonify({"ok": True})


# ──────────────────────────────────────────────────────────────────────
#  Team meeting notes (personal notes recorded by team members)
# ──────────────────────────────────────────────────────────────────────
def _can_access_team_notes(u, team):
    """admin/analyst can access any team; team-role users only their own team(s)."""
    if not u:
        return False
    if u["role"] in ("admin", "analyst"):
        return True
    if u["role"] == "team" and u["team"]:
        return team in parse_teams(u["team"])
    return False


@app.route("/api/team-meeting-notes", methods=["GET"])
@login_required
def team_meeting_notes_list():
    u = get_current_user()
    team = request.args.get("team", "")
    if u["role"] == "team" and u["team"]:
        team = u["team"]
    with closing(get_db()) as db:
        q = "SELECT * FROM team_meeting_notes WHERE 1=1"
        params = []
        if team:
            team_set = parse_teams(team)
            placeholders = ",".join("?" * len(team_set))
            q += f" AND team IN ({placeholders})"
            params.extend(sorted(team_set))
        rows = db.execute(q + " ORDER BY note_date DESC, id DESC", params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/team-meeting-notes", methods=["POST"])
@login_required
def team_meeting_notes_create():
    u = get_current_user()
    data = request.get_json(force=True, silent=True) or {}
    team = str(data.get("team", "")).strip()
    note_date = str(data.get("note_date", "")).strip()
    title = str(data.get("title", "")).strip()
    notes = str(data.get("notes", "")).strip()
    if not team or not note_date:
        return jsonify({"error": "team and note_date required"}), 400
    if not _can_access_team_notes(u, team):
        return jsonify({"error": "forbidden"}), 403
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        db.execute(
            "INSERT INTO team_meeting_notes (team,note_date,title,notes,created_by,created_at,updated_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (team, note_date, title, notes, u["username"], now, now)
        )
        db.commit()
        nid = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    return jsonify({"ok": True, "id": nid}), 201


@app.route("/api/team-meeting-notes/<int:nid>", methods=["PUT"])
@login_required
def team_meeting_notes_update(nid):
    u = get_current_user()
    with closing(get_db()) as db:
        existing = db.execute("SELECT * FROM team_meeting_notes WHERE id=?", (nid,)).fetchone()
    if not existing:
        return jsonify({"error": "not found"}), 404
    existing = dict(existing)
    if not _can_access_team_notes(u, existing["team"]):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json(force=True, silent=True) or {}
    note_date = str(data.get("note_date", existing["note_date"])).strip()
    title = str(data.get("title", existing["title"] or "")).strip()
    notes = str(data.get("notes", existing["notes"] or "")).strip()
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        db.execute(
            "UPDATE team_meeting_notes SET note_date=?,title=?,notes=?,updated_at=? WHERE id=?",
            (note_date, title, notes, now, nid)
        )
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/team-meeting-notes/<int:nid>", methods=["DELETE"])
@login_required
def team_meeting_notes_delete(nid):
    u = get_current_user()
    with closing(get_db()) as db:
        existing = db.execute("SELECT * FROM team_meeting_notes WHERE id=?", (nid,)).fetchone()
    if not existing:
        return jsonify({"error": "not found"}), 404
    existing = dict(existing)
    if not _can_access_team_notes(u, existing["team"]):
        return jsonify({"error": "forbidden"}), 403
    with closing(get_db()) as db:
        db.execute("DELETE FROM team_meeting_notes WHERE id=?", (nid,))
        db.commit()
    return jsonify({"ok": True})


# ──────────────────────────────────────────────────────────────────────
#  Dashboard routes
# ──────────────────────────────────────────────────────────────────────
@app.route("/api/health")
def health():
    return jsonify({"status": "ok"})


@app.route("/api/status")
@login_required
def status():
    out = {}
    for k in PARSERS:
        out[k] = dataset_meta(k)
    return jsonify(out)


@app.route("/api/upload/<key>", methods=["POST"])
@admin_required
def upload(key):
    if key not in PARSERS:
        return jsonify({"error": "unknown dataset key"}), 400
    if "file" not in request.files:
        return jsonify({"error": "no file"}), 400
    f = request.files["file"]
    try:
        parser, _ = PARSERS[key]
        if key in TEXT_PARSERS:
            records = parser(f.read(), f.filename)
        else:
            sheets = read_excel_any(f)
            records = parser(sheets)
        if not records:
            return jsonify({"error": "No valid rows parsed. Check the file format."}), 400
        save_dataset(key, f.filename, records)
        return jsonify({"ok": True, "key": key, "filename": f.filename,
                        "rowcount": len(records)})
    except Exception as e:  # noqa
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/clear/<key>", methods=["POST"])
@admin_required
def clear(key):
    clear_key(key)
    return jsonify({"ok": True})


@app.route("/api/clear-all", methods=["POST"])
@admin_required
def clear_all():
    for k in PARSERS:
        clear_key(k)
    return jsonify({"ok": True})


@app.route("/api/dashboard", methods=["POST"])
@login_required
def dashboard():
    u = get_current_user()
    filters = request.get_json(force=True, silent=True) or {}
    filters["team"] = resolve_team_filter(u, filters.get("team"))
    return jsonify(compute(filters))


@app.route("/api/po-detail", methods=["GET"])
@login_required
def po_detail():
    """Return raw open-PO rows (for the open-PO table)."""
    u = get_current_user()
    team_filter = request.args.get("team", "")
    if u and u["role"] == "team" and u["team"]:
        team_filter = u["team"]
    po = build_po_enriched()
    if team_filter:
        team_set = parse_teams(team_filter)
        po = [p for p in po if p["team"] in team_set]
    po.sort(key=lambda p: (p.get("ship_date") or "9999-12-31"))
    return jsonify(po[:500])  # cap


@app.route("/api/program-margin", methods=["POST"])
@login_required
def program_margin():
    u = get_current_user()
    filters = request.get_json(force=True, silent=True) or {}
    f_by = filters.get("by")
    f_team = parse_teams(resolve_team_filter(u, filters.get("team")))
    f_cust = filters.get("customer") or ""
    f_cat = filters.get("category") or ""

    zecost_ds = load_dataset("zecost06")
    if not zecost_ds:
        return jsonify({"error": "Cost data (zecost06) not uploaded"})

    sales_ds = load_dataset("sales")
    po_ds = load_dataset("po")
    prod_cat_ds = load_dataset("prod_cat_master")
    zexpr_ds = load_dataset("zexpr24")
    mis_map = get_mis_map()
    cust_map = get_cust_map()

    # Build export_id → to_date from program master (take latest to_date per eid).
    # Index by normalised key (leading zeros stripped) so that IDs like
    # "0000201257LGP2025" (zecost06) match "201257LGP2025" (zexpr24) and vice-versa.
    today = dt.date.today()
    eid_to_date = {}  # normalised_eid → to_date

    def _norm_eid(e):
        s = str(e).strip()
        return s.lstrip("0") or s  # preserve "0" if entirely zeros

    if zexpr_ds:
        for r in zexpr_ds["records"]:
            raw = r.get("export_id", "").strip()
            if not raw:
                continue
            key = _norm_eid(raw)
            to_d = None
            if r.get("to_y") and r.get("to_m") and r.get("to_d_day"):
                try:
                    to_d = dt.date(r["to_y"], r["to_m"], r["to_d_day"])
                except Exception:
                    pass
            if key not in eid_to_date or (to_d and (eid_to_date[key] is None or to_d > eid_to_date[key])):
                eid_to_date[key] = to_d

    # Determine BY from sales if not supplied
    if not f_by and sales_ds:
        all_bys_s = sorted(
            {business_year(dt.date(r["y"], r["m"], r["d"]))
             for r in sales_ds["records"] if r.get("y")},
            reverse=True)
        f_by = all_bys_s[0] if all_bys_s else None
    if not f_by:
        return jsonify({"error": "No business year available"})

    pby = prior_by(f_by)
    a_s, b_s = f_by.split("-")
    by_start = dt.date(int(a_s), 10, 1)
    by_end = dt.date(int("20" + b_s), 9, 30)

    # Build prod_cat lookup as fallback: material → (third, sub, cat)
    prod_cat_map = {}
    if prod_cat_ds:
        for r in prod_cat_ds["records"]:
            prod_cat_map[r["material"]] = (r.get("third", ""), r.get("sub", ""), r.get("cat", ""))

    rows = []
    for r in zecost_ds["records"]:
        eid = r.get("export_id", "").strip()
        if not eid:
            continue

        # Use MIS category stored in the cost file; fall back to product category master
        mis_cat = r.get("mis_cat", "").strip()
        if not mis_cat:
            third = r.get("third", "")
            sub   = r.get("sub", "")
            cat   = r.get("cat", "")
            if not any([third, sub, cat]):
                mat = r["material"]
                third, sub, cat = prod_cat_map.get(mat, ("", "", ""))
            mis_cat = resolve_mis_category(third, sub, cat, mis_map) or "Uncategorised"

        cm = cust_map.get(r["customer_code"], {})
        team = cm.get("team") or "—"
        mis_cust = cm.get("mis_customer") or cm.get("name") or r["customer_code"] or "—"

        if f_team and team not in f_team:
            continue
        if f_cust and mis_cust != f_cust:
            continue
        if f_cat and mis_cat != f_cat:
            continue

        rows.append({
            "export_id": eid,
            "team": team, "mis_customer": mis_cust, "mis_cat": mis_cat,
            "po_cost": r["po_cost_usd"], "po_sale": r["po_sale_usd"],
            "quot_cost": r["quot_cost_usd"], "quot_sale": r["quot_sale_usd"],
        })

    def agg(items, keyfn):
        g = {}
        for item in items:
            k = keyfn(item)
            if k not in g:
                g[k] = {"po_cost": 0.0, "po_sale": 0.0, "quot_cost": 0.0, "quot_sale": 0.0}
            for fld in ("po_cost", "po_sale", "quot_cost", "quot_sale"):
                g[k][fld] += item[fld]
        return g

    def mgn(sale, cost):
        return round((sale - cost) / sale * 100, 1) if sale != 0 else None

    def serialize_mg(g):
        out = []
        for k, v in g.items():
            pm = mgn(v["po_sale"], v["po_cost"])
            qm = mgn(v["quot_sale"], v["quot_cost"])
            out.append({
                "name": k or "—",
                "po_sale_mn": round(v["po_sale"] / 1_000_000, 3),
                "quot_sale_mn": round(v["quot_sale"] / 1_000_000, 3),
                "po_margin": pm,
                "quot_margin": qm,
                "delta": round((pm or 0) - (qm or 0), 1)
                         if pm is not None and qm is not None else None,
            })
        out.sort(key=lambda x: abs(x["po_sale_mn"]) + abs(x["quot_sale_mn"]), reverse=True)
        return out

    by_cust = agg(rows, lambda r: r["mis_customer"])
    by_cat = agg(rows, lambda r: r["mis_cat"])

    t_ps = sum(r["po_sale"] for r in rows)
    t_pc = sum(r["po_cost"] for r in rows)
    t_qs = sum(r["quot_sale"] for r in rows)
    t_qc = sum(r["quot_cost"] for r in rows)

    # Aggregate by export_id for program-level detail
    by_prog = {}
    for r in rows:
        eid = r["export_id"]
        if eid not in by_prog:
            to_d = eid_to_date.get(_norm_eid(eid))
            is_open = to_d is None or to_d >= today
            by_prog[eid] = {
                "export_id": eid,
                "customer": r["mis_customer"],
                "to_date": to_d.isoformat() if to_d else None,
                "is_open": is_open,
                "po_cost": 0.0, "po_sale": 0.0,
                "quot_cost": 0.0, "quot_sale": 0.0,
            }
        for fld in ("po_cost", "po_sale", "quot_cost", "quot_sale"):
            by_prog[eid][fld] += r[fld]

    def serialize_progs(prog_dict):
        out = []
        for v in prog_dict.values():
            pm = mgn(v["po_sale"], v["po_cost"])
            qm = mgn(v["quot_sale"], v["quot_cost"])
            out.append({
                "export_id": v["export_id"],
                "customer": v["customer"],
                "to_date": v["to_date"],
                "is_open": v["is_open"],
                "po_sale_mn":   round(v["po_sale"]   / 1_000_000, 3),
                "quot_sale_mn": round(v["quot_sale"] / 1_000_000, 3),
                "po_margin":   pm,
                "quot_margin": qm,
                "delta": round((pm or 0) - (qm or 0), 1)
                         if pm is not None and qm is not None else None,
            })
        out.sort(key=lambda x: (x["customer"].lower(), -(abs(x["po_sale_mn"]) + abs(x["quot_sale_mn"]))))
        return out

    prog_list = serialize_progs(by_prog)
    open_progs   = [p for p in prog_list if p["is_open"]]
    closed_progs = [p for p in prog_list if not p["is_open"]]

    all_bys = []
    if sales_ds:
        all_bys = sorted(
            {business_year(dt.date(r["y"], r["m"], r["d"]))
             for r in sales_ds["records"] if r.get("y")},
            reverse=True)

    return jsonify({
        "by": f_by, "prior_by": pby, "all_bys": all_bys,
        "active_programs": len({r["export_id"] for r in rows}),
        "by_customer": serialize_mg(by_cust),
        "by_category": serialize_mg(by_cat),
        "open_programs":   open_progs,
        "closed_programs": closed_progs,
        "totals": {
            "po_sale_mn": round(t_ps / 1_000_000, 3),
            "quot_sale_mn": round(t_qs / 1_000_000, 3),
            "po_margin": mgn(t_ps, t_pc),
            "quot_margin": mgn(t_qs, t_qc),
            "delta": round((mgn(t_ps, t_pc) or 0) - (mgn(t_qs, t_qc) or 0), 1)
                     if t_ps and t_qs else None,
        },
        "options": {
            "teams": sorted({r["team"] for r in rows if r["team"] != "—"}),
            "customers": sorted({r["mis_customer"] for r in rows if r["mis_customer"] != "—"}),
            "categories": sorted({r["mis_cat"] for r in rows if r["mis_cat"]}),
        }
    })


# ──────────────────────────────────────────────────────────────────────
#  Market Intelligence — public read endpoints
# ──────────────────────────────────────────────────────────────────────
@app.route("/api/mi/categories", methods=["GET"])
@login_required
def mi_categories():
    """Active categories with their products — used by frontend filters."""
    with closing(get_db()) as db:
        cats = db.execute(
            "SELECT * FROM mi_categories ORDER BY sort_order, id"
        ).fetchall()
        result = []
        for cat in cats:
            cd = dict(cat)
            for f in ("keywords", "hs_codes"):
                try:
                    cd[f] = json.loads(cd.get(f) or "[]")
                except Exception:
                    cd[f] = []
            prods = db.execute(
                "SELECT * FROM mi_products WHERE category_id=? ORDER BY sort_order, id",
                (cd["id"],)
            ).fetchall()
            pd_list = []
            for p in prods:
                pd = dict(p)
                for f in ("keywords", "alternate_names", "hs_codes"):
                    try:
                        pd[f] = json.loads(pd.get(f) or "[]")
                    except Exception:
                        pd[f] = []
                pd_list.append(pd)
            cd["products"] = pd_list
            result.append(cd)
    return jsonify(result)


@app.route("/api/mi/team-markets", methods=["GET"])
@login_required
def mi_team_markets():
    with closing(get_db()) as db:
        rows = db.execute(
            "SELECT * FROM mi_team_markets ORDER BY team_name, sort_order"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/mi/regions", methods=["GET"])
@login_required
def mi_regions():
    """Return {region: [country, ...]} built from active mi_team_markets rows
    that have a region tagged."""
    with closing(get_db()) as db:
        rows = db.execute(
            "SELECT DISTINCT region, market FROM mi_team_markets "
            "WHERE is_active=1 AND region!='' ORDER BY region, market"
        ).fetchall()
    out = {}
    for r in rows:
        out.setdefault(r["region"], [])
        if r["market"] not in out[r["region"]]:
            out[r["region"]].append(r["market"])
    return jsonify(out)


@app.route("/api/mi/updates", methods=["GET"])
@login_required
def mi_updates():
    """Returns intelligence updates with optional filters."""
    u = get_current_user()
    month  = request.args.get("month", type=int)
    year   = request.args.get("year",  type=int)
    team   = request.args.get("team",  "")
    market = request.args.get("market","")
    cat    = request.args.get("category","")
    src_t  = request.args.get("source_type","")
    search = request.args.get("search","").strip().lower()

    # Team users see only their own team
    if u and u["role"] == "team" and u["team"]:
        team = u["team"]

    # Default to current month/year if not specified
    today = dt.date.today()
    if not year:  year  = today.year
    if not month: month = today.month

    with closing(get_db()) as db:
        q = "SELECT * FROM mi_updates WHERE year=? AND month=?"
        params = [year, month]
        if team:
            team_set = parse_teams(team)
            placeholders = ",".join("?" * len(team_set))
            q += f" AND team_name IN ({placeholders})"
            params.extend(sorted(team_set))
        if market: q += " AND market=?";         params.append(market)
        if cat:    q += " AND category_name=?";  params.append(cat)
        if src_t:  q += " AND source_type=?";    params.append(src_t)
        rows = db.execute(q + " ORDER BY credibility_score DESC, relevance_score DESC", params).fetchall()

        # Get available months for dropdown
        months = db.execute(
            "SELECT DISTINCT year, month FROM mi_updates ORDER BY year DESC, month DESC LIMIT 24"
        ).fetchall()

        # Get last job completion
        last_log = db.execute(
            "SELECT completed_at FROM mi_job_logs WHERE status='completed' ORDER BY id DESC LIMIT 1"
        ).fetchone()

    items = [dict(r) for r in rows]
    # Apply text search client-side filter (server-side for simplicity)
    if search:
        items = [r for r in items if search in (r["update_title"] + r["update_summary"] + r["market"]).lower()]

    return jsonify({
        "updates": items,
        "month": month,
        "year": year,
        "last_run": last_log["completed_at"] if last_log else None,
        "available_months": [{"year": r["year"], "month": r["month"]} for r in months],
    })


# ──────────────────────────────────────────────────────────────────────
#  Market Intelligence — admin configuration endpoints
# ──────────────────────────────────────────────────────────────────────
@app.route("/api/admin/mi/status", methods=["GET"])
@admin_required
def mi_admin_status():
    import os as _os
    return jsonify({
        "has_nvidia":     bool(_os.environ.get("NVIDIA_API_KEY")),
        "has_openrouter": bool(_os.environ.get("OPENROUTER_API_KEY")),
        "has_tavily":     bool(_os.environ.get("TAVILY_API_KEY")),
        "model":          _os.environ.get("NVIDIA_MODEL", "nvidia/nemotron-3-ultra-550b-a55b"),
        "fallback_model": _os.environ.get("MI_MODEL", "openai/gpt-oss-20b:free"),
    })




@app.route("/api/admin/mi/categories", methods=["POST"])
@admin_required
def mi_cat_create():
    data = request.get_json(force=True, silent=True) or {}
    name = str(data.get("name", "")).strip()
    if not name:
        return jsonify({"error": "name required"}), 400
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        db.execute(
            "INSERT INTO mi_categories (name,description,keywords,hs_codes,is_active,sort_order,created_at,updated_at) VALUES(?,?,?,?,1,?,?,?)",
            (name, str(data.get("description","")).strip(),
             json.dumps(data.get("keywords") or []),
             json.dumps(data.get("hs_codes") or []),
             int(data.get("sort_order") or 0), now, now)
        )
        db.commit()
        nid = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    return jsonify({"ok": True, "id": nid}), 201


@app.route("/api/admin/mi/categories/<int:cid>", methods=["PUT"])
@admin_required
def mi_cat_update(cid):
    data = request.get_json(force=True, silent=True) or {}
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        db.execute(
            "UPDATE mi_categories SET name=?,description=?,keywords=?,hs_codes=?,is_active=?,sort_order=?,updated_at=? WHERE id=?",
            (str(data.get("name","")).strip(), str(data.get("description","")).strip(),
             json.dumps(data.get("keywords") or []),
             json.dumps(data.get("hs_codes") or []),
             1 if data.get("is_active", True) else 0,
             int(data.get("sort_order") or 0), now, cid)
        )
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/mi/categories/<int:cid>", methods=["DELETE"])
@admin_required
def mi_cat_delete(cid):
    with closing(get_db()) as db:
        # Soft-delete: deactivate
        db.execute("UPDATE mi_categories SET is_active=0 WHERE id=?", (cid,))
        db.execute("UPDATE mi_products SET is_active=0 WHERE category_id=?", (cid,))
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/mi/categories/download", methods=["GET"])
@admin_required
def mi_cat_download():
    """Download all active categories and products as a two-sheet Excel file."""
    with closing(get_db()) as db:
        cats = db.execute(
            "SELECT id, name, description, keywords, hs_codes, sort_order FROM mi_categories WHERE is_active=1 ORDER BY sort_order, id"
        ).fetchall()
        prods = db.execute(
            "SELECT p.name AS product, c.name AS category, p.keywords, p.hs_codes, p.sort_order "
            "FROM mi_products p JOIN mi_categories c ON c.id=p.category_id "
            "WHERE p.is_active=1 AND c.is_active=1 ORDER BY c.sort_order, c.id, p.sort_order, p.id"
        ).fetchall()

    cat_rows = []
    for c in cats:
        try:
            kw = ", ".join(json.loads(c["keywords"] or "[]"))
        except Exception:
            kw = ""
        try:
            hs = ", ".join(json.loads(c["hs_codes"] or "[]"))
        except Exception:
            hs = ""
        cat_rows.append({
            "Category Name":  c["name"],
            "Description":    c["description"] or "",
            "Keywords":       kw,
            "HS Codes":       hs,
            "Sort Order":     c["sort_order"],
        })

    prod_rows = []
    for p in prods:
        try:
            kw = ", ".join(json.loads(p["keywords"] or "[]"))
        except Exception:
            kw = ""
        try:
            hs = ", ".join(json.loads(p["hs_codes"] or "[]"))
        except Exception:
            hs = ""
        prod_rows.append({
            "Category Name":  p["category"],
            "Product Name":   p["product"],
            "Keywords":       kw,
            "HS Codes":       hs,
            "Sort Order":     p["sort_order"],
        })

    df_cats  = pd.DataFrame(cat_rows)  if cat_rows  else pd.DataFrame(columns=["Category Name","Description","Keywords","HS Codes","Sort Order"])
    df_prods = pd.DataFrame(prod_rows) if prod_rows else pd.DataFrame(columns=["Category Name","Product Name","Keywords","HS Codes","Sort Order"])

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df_cats.to_excel(writer, index=False, sheet_name="Categories")
        df_prods.to_excel(writer, index=False, sheet_name="Products")
        for sheet in writer.sheets.values():
            for col_cells in sheet.columns:
                width = max((len(str(c.value or "")) for c in col_cells), default=8) + 4
                sheet.column_dimensions[col_cells[0].column_letter].width = min(width, 60)
    bio.seek(0)

    from datetime import date as _date
    fname = f"product_categories_{_date.today().isoformat()}.xlsx"
    return send_file(bio,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=fname)


@app.route("/api/admin/mi/categories/upload", methods=["POST"])
@admin_required
def mi_cat_upload():
    """Upload Excel to upsert categories and products (add new, update existing by name)."""
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    raw = f.read()
    bio = io.BytesIO(raw)
    try:
        xls = pd.read_excel(bio, sheet_name=None, dtype=str)
    except Exception as exc:
        return jsonify({"error": f"Cannot read file: {exc}"}), 400

    now = dt.datetime.now().isoformat(timespec="seconds")
    stats = {"categories_added": 0, "categories_updated": 0,
             "products_added": 0, "products_updated": 0, "errors": []}

    def _parse_list(val):
        """'a, b, c' → ['a','b','c']"""
        if not val or str(val).strip() in ("", "nan"):
            return []
        return [x.strip() for x in str(val).split(",") if x.strip()]

    with closing(get_db()) as db:
        # ── Categories sheet ──────────────────────────────────────────────
        cat_sheet = xls.get("Categories") or xls.get("categories")
        if cat_sheet is not None:
            cat_sheet.columns = [str(c).strip() for c in cat_sheet.columns]
            for i, row in cat_sheet.iterrows():
                name = str(row.get("Category Name", "")).strip()
                if not name or name == "nan":
                    continue
                desc     = str(row.get("Description", "") or "").strip()
                keywords = json.dumps(_parse_list(row.get("Keywords", "")))
                hs_codes = json.dumps(_parse_list(row.get("HS Codes", "")))
                try:
                    sort_order = int(float(str(row.get("Sort Order", 0) or 0)))
                except Exception:
                    sort_order = 0
                existing = db.execute(
                    "SELECT id FROM mi_categories WHERE name=?", (name,)
                ).fetchone()
                if existing:
                    db.execute(
                        "UPDATE mi_categories SET description=?,keywords=?,hs_codes=?,sort_order=?,is_active=1,updated_at=? WHERE id=?",
                        (desc, keywords, hs_codes, sort_order, now, existing["id"])
                    )
                    stats["categories_updated"] += 1
                else:
                    db.execute(
                        "INSERT INTO mi_categories (name,description,keywords,hs_codes,is_active,sort_order,created_at,updated_at) VALUES(?,?,?,?,1,?,?,?)",
                        (name, desc, keywords, hs_codes, sort_order, now, now)
                    )
                    stats["categories_added"] += 1

        # ── Products sheet ────────────────────────────────────────────────
        prod_sheet = xls.get("Products") or xls.get("products")
        if prod_sheet is not None:
            prod_sheet.columns = [str(c).strip() for c in prod_sheet.columns]
            for i, row in prod_sheet.iterrows():
                cat_name  = str(row.get("Category Name", "")).strip()
                prod_name = str(row.get("Product Name", "")).strip()
                if not cat_name or not prod_name or cat_name == "nan" or prod_name == "nan":
                    continue
                keywords = json.dumps(_parse_list(row.get("Keywords", "")))
                hs_codes = json.dumps(_parse_list(row.get("HS Codes", "")))
                try:
                    sort_order = int(float(str(row.get("Sort Order", 0) or 0)))
                except Exception:
                    sort_order = 0
                cat_row = db.execute(
                    "SELECT id FROM mi_categories WHERE name=? AND is_active=1", (cat_name,)
                ).fetchone()
                if not cat_row:
                    stats["errors"].append(f"Row {i+2}: category '{cat_name}' not found — add it in the Categories sheet first.")
                    continue
                cat_id = cat_row["id"]
                existing = db.execute(
                    "SELECT id FROM mi_products WHERE category_id=? AND name=?", (cat_id, prod_name)
                ).fetchone()
                if existing:
                    db.execute(
                        "UPDATE mi_products SET keywords=?,hs_codes=?,sort_order=?,is_active=1,updated_at=? WHERE id=?",
                        (keywords, hs_codes, sort_order, now, existing["id"])
                    )
                    stats["products_updated"] += 1
                else:
                    db.execute(
                        "INSERT INTO mi_products (category_id,name,keywords,alternate_names,hs_codes,is_active,sort_order,created_at,updated_at) VALUES(?,?,?,?,?,1,?,?,?)",
                        (cat_id, prod_name, keywords, json.dumps([]), hs_codes, sort_order, now, now)
                    )
                    stats["products_added"] += 1

        db.commit()

    return jsonify({"ok": True, **stats})


@app.route("/api/admin/mi/products", methods=["POST"])
@admin_required
def mi_prod_create():
    data = request.get_json(force=True, silent=True) or {}
    name   = str(data.get("name", "")).strip()
    cat_id = data.get("category_id")
    if not name or not cat_id:
        return jsonify({"error": "name and category_id required"}), 400
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        db.execute(
            "INSERT INTO mi_products (category_id,name,description,keywords,alternate_names,hs_codes,is_active,sort_order,created_at,updated_at) VALUES(?,?,?,?,?,?,1,?,?,?)",
            (int(cat_id), name, str(data.get("description","")).strip(),
             json.dumps(data.get("keywords") or []),
             json.dumps(data.get("alternate_names") or []),
             json.dumps(data.get("hs_codes") or []),
             int(data.get("sort_order") or 0), now, now)
        )
        db.commit()
        nid = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    return jsonify({"ok": True, "id": nid}), 201


@app.route("/api/admin/mi/products/<int:pid>", methods=["PUT"])
@admin_required
def mi_prod_update(pid):
    data = request.get_json(force=True, silent=True) or {}
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        db.execute(
            "UPDATE mi_products SET name=?,description=?,keywords=?,alternate_names=?,hs_codes=?,is_active=?,sort_order=?,updated_at=? WHERE id=?",
            (str(data.get("name","")).strip(), str(data.get("description","")).strip(),
             json.dumps(data.get("keywords") or []),
             json.dumps(data.get("alternate_names") or []),
             json.dumps(data.get("hs_codes") or []),
             1 if data.get("is_active", True) else 0,
             int(data.get("sort_order") or 0), now, pid)
        )
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/mi/products/<int:pid>", methods=["DELETE"])
@admin_required
def mi_prod_delete(pid):
    with closing(get_db()) as db:
        db.execute("UPDATE mi_products SET is_active=0 WHERE id=?", (pid,))
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/mi/team-markets", methods=["POST"])
@admin_required
def mi_tm_create():
    data = request.get_json(force=True, silent=True) or {}
    team   = str(data.get("team_name","")).strip()
    market = str(data.get("market","")).strip()
    if not team or not market:
        return jsonify({"error": "team_name and market required"}), 400
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        db.execute(
            "INSERT INTO mi_team_markets (team_name,market,country_code,region,is_active,sort_order,created_at) VALUES(?,?,?,?,1,?,?)",
            (team, market, str(data.get("country_code","")).strip().upper(),
             str(data.get("region","")).strip(),
             int(data.get("sort_order") or 0), now)
        )
        db.commit()
        nid = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    return jsonify({"ok": True, "id": nid}), 201


@app.route("/api/admin/mi/team-markets/<int:tid>", methods=["PUT"])
@admin_required
def mi_tm_update(tid):
    data = request.get_json(force=True, silent=True) or {}
    with closing(get_db()) as db:
        db.execute(
            "UPDATE mi_team_markets SET team_name=?,market=?,country_code=?,region=?,is_active=?,sort_order=? WHERE id=?",
            (str(data.get("team_name","")).strip(), str(data.get("market","")).strip(),
             str(data.get("country_code","")).strip().upper(),
             str(data.get("region","")).strip(),
             1 if data.get("is_active", True) else 0,
             int(data.get("sort_order") or 0), tid)
        )
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/mi/team-markets/<int:tid>", methods=["DELETE"])
@admin_required
def mi_tm_delete(tid):
    with closing(get_db()) as db:
        db.execute("DELETE FROM mi_team_markets WHERE id=?", (tid,))
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/mi/team-markets/export", methods=["GET"])
@admin_required
def mi_tm_export():
    """Download all team-market mappings as Excel."""
    import openpyxl
    from io import BytesIO
    with closing(get_db()) as db:
        rows = db.execute(
            "SELECT team_name, market, country_code, region, is_active, sort_order "
            "FROM mi_team_markets ORDER BY team_name, sort_order"
        ).fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Team-Market Mapping"
    ws.append(["Team Name", "Market / Country", "Country Code", "Region", "Active (1=Yes 0=No)", "Sort Order"])
    for r in rows:
        ws.append([r["team_name"], r["market"], r["country_code"], r["region"] or "",
                   1 if r["is_active"] else 0, r["sort_order"] or 0])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="team_market_mapping.xlsx")


@app.route("/api/admin/mi/team-markets/import", methods=["POST"])
@admin_required
def mi_tm_import():
    """Upload Excel to bulk-upsert team-market mappings."""
    import openpyxl
    from io import BytesIO
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({"error": "Empty file"}), 400
        # Auto-detect header row
        header = [str(c).strip().lower().replace(" ", "").replace("/", "").replace("(", "").replace(")", "") if c else "" for c in rows[0]]
        def col_idx(*names):
            for n in names:
                nn = n.lower().replace(" ", "").replace("/", "")
                for i, h in enumerate(header):
                    if nn in h:
                        return i
            return None
        ti = col_idx("teamname", "team")
        mi = col_idx("market", "country")
        ci = col_idx("countrycode", "code")
        ri = col_idx("region")
        ai = col_idx("active")
        oi = col_idx("sortorder", "sort")
        if ti is None or mi is None:
            return jsonify({"error": "Could not find Team Name or Market columns"}), 400
        now = dt.datetime.now().isoformat(timespec="seconds")
        inserted = 0
        with closing(get_db()) as db:
            for row in rows[1:]:
                team = str(row[ti]).strip() if row[ti] is not None else ""
                market = str(row[mi]).strip() if row[mi] is not None else ""
                if not team or not market or team.lower() in ("none", "team name"):
                    continue
                code = str(row[ci]).strip().upper() if ci is not None and row[ci] is not None else ""
                region = str(row[ri]).strip() if ri is not None and row[ri] is not None else ""
                active = 1
                if ai is not None and row[ai] is not None:
                    av = str(row[ai]).strip().lower()
                    active = 0 if av in ("0", "no", "false", "inactive") else 1
                sort = 0
                if oi is not None and row[oi] is not None:
                    try: sort = int(float(str(row[oi])))
                    except: pass
                existing = db.execute(
                    "SELECT id FROM mi_team_markets WHERE team_name=? AND market=?",
                    (team, market)
                ).fetchone()
                if existing:
                    db.execute(
                        "UPDATE mi_team_markets SET country_code=?,region=?,is_active=?,sort_order=? WHERE id=?",
                        (code, region, active, sort, existing["id"])
                    )
                else:
                    db.execute(
                        "INSERT INTO mi_team_markets (team_name,market,country_code,region,is_active,sort_order,created_at) VALUES(?,?,?,?,?,?,?)",
                        (team, market, code, region, active, sort, now)
                    )
                inserted += 1
            db.commit()
        return jsonify({"ok": True, "imported": inserted})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/admin/mi/updates/<int:uid>", methods=["DELETE"])
@admin_required
def mi_update_delete(uid):
    with closing(get_db()) as db:
        db.execute("DELETE FROM mi_updates WHERE id=?", (uid,))
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/mi/job-logs", methods=["GET"])
@admin_required
def mi_job_logs():
    with closing(get_db()) as db:
        rows = db.execute(
            "SELECT * FROM mi_job_logs ORDER BY id DESC LIMIT 30"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


def _mi_missing_keys():
    """Required env vars for Market Intelligence / Market Info AI jobs:
    TAVILY_API_KEY plus at least one of NVIDIA_API_KEY / OPENROUTER_API_KEY."""
    missing = []
    if not (os.environ.get("NVIDIA_API_KEY") or os.environ.get("OPENROUTER_API_KEY")):
        missing.append("NVIDIA_API_KEY or OPENROUTER_API_KEY")
    if not os.environ.get("TAVILY_API_KEY"):
        missing.append("TAVILY_API_KEY")
    return missing


@app.route("/api/admin/mi/run-job", methods=["POST"])
@admin_required
def mi_run_job():
    """Trigger the intelligence job asynchronously."""
    try:
        import intelligence as _mi
    except ImportError:
        return jsonify({"error": "intelligence module not found"}), 500

    # Preflight: need TAVILY + at least one AI provider key before spawning the thread
    missing = _mi_missing_keys()
    if missing:
        return jsonify({
            "error": f"Cannot run job — missing environment variable(s): {', '.join(missing)}. "
                     "Add them to docker-compose.yml and restart."
        }), 400

    data = request.get_json(force=True, silent=True) or {}
    month = data.get("month") or dt.date.today().month
    year  = data.get("year")  or dt.date.today().year
    teams_filter = data.get("teams") or None

    def _run():
        try:
            _mi.run_intelligence_job(month=month, year=year, teams_filter=teams_filter)
        except Exception as exc:
            logger.error("Background MI job error: %s", exc)

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return jsonify({"ok": True, "message": f"Job started for {month}/{year} — check job logs for progress."})


# ══════════════════════════════════════════════════════════════════════
# Market Information endpoints
# ══════════════════════════════════════════════════════════════════════

@app.route("/api/market-info/freshness", methods=["GET"])
@login_required
def market_info_freshness():
    country  = request.args.get("country", "").strip()
    category = request.args.get("category", "").strip()
    if not country or not category:
        return jsonify({"error": "country and category required"}), 400
    try:
        import market_info as _mi
        rec = _mi.get_cached(country, category)
        if not rec:
            return jsonify({"status": "missing", "fresh": False})
        fresh = _mi.is_fresh(rec)
        throttled = _mi.was_refreshed_today(country, category)
        return jsonify({
            "status": rec.get("cache_status", "ok"),
            "fresh": fresh,
            "fetch_date": rec.get("fetch_date"),
            "next_update": rec.get("next_update"),
            "attempt_count": rec.get("attempt_count", 0),
            "throttled": throttled,
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


def _get_category_meta(category: str) -> dict:
    """Return hs_codes and active product names for a category."""
    with closing(get_db()) as db:
        cat_row = db.execute(
            "SELECT id, hs_codes FROM mi_categories WHERE name=? AND is_active=1 LIMIT 1",
            (category,)
        ).fetchone()
        if not cat_row:
            return {"hs_codes": [], "products": []}
        try:
            hs = json.loads(cat_row["hs_codes"] or "[]") or []
        except Exception:
            hs = []
        prod_rows = db.execute(
            "SELECT name FROM mi_products WHERE category_id=? AND is_active=1 ORDER BY sort_order, id",
            (cat_row["id"],)
        ).fetchall()
    return {"hs_codes": hs, "products": [r["name"] for r in prod_rows]}


def _market_info_cached_data(rec: dict | None) -> dict | None:
    if not rec:
        return None
    raw = (rec.get("data_json") or "").strip()
    if not raw:
        return None
    try:
        data = json.loads(raw)
    except Exception as exc:
        logger.warning(
            "Ignoring invalid market-info cache for %s/%s: %s",
            rec.get("country"), rec.get("category"), exc,
        )
        return None
    return data if isinstance(data, dict) and data else None


def _market_info_error_message(exc: Exception) -> str:
    text = str(exc)
    low = text.lower()
    if "429" in text or "too many requests" in low or "rate-limit" in low or "rate limited" in low:
        return (
            "OpenRouter is rate-limited right now. Showing cached data where available; "
            "wait a few minutes before refreshing uncached categories."
        )
    if "invalid json" in low or "empty response" in low or "expecting value" in low:
        return "AI response was not valid JSON. Showing cached data where available."
    return text


@app.route("/api/market-info", methods=["GET"])
@login_required
def market_info_get():
    country  = request.args.get("country", "").strip()
    category = request.args.get("category", "").strip()
    if not country or not category:
        return jsonify({"error": "country and category required"}), 400
    try:
        import market_info as _mi
        rec = _mi.get_cached(country, category)
        if rec and _mi.is_fresh(rec) and rec.get("cache_status") == "ok":
            data = _market_info_cached_data(rec)
            if data:
                return jsonify({
                    "ok": True,
                    "from_cache": True,
                    "fetch_date": rec["fetch_date"],
                    "next_update": rec["next_update"],
                    "data": data,
                })
        # Stale or missing — fetch now
        missing = _mi_missing_keys()
        if missing:
            data = _market_info_cached_data(rec)
            if data:
                return jsonify({
                    "ok": True, "from_cache": True, "stale": True,
                    "fetch_date": rec["fetch_date"],
                    "next_update": rec["next_update"],
                    "data": data,
                })
            return jsonify({"error": f"API keys missing: {', '.join(missing)}"}), 503
        _cat_meta = _get_category_meta(category)
        data = _mi.fetch_market_info(country, category, _cat_meta["hs_codes"], _cat_meta["products"], fast=True)
        _mi.save_cache(country, category, data, triggered_by="on_demand")
        rec2 = _mi.get_cached(country, category)
        return jsonify({
            "ok": True,
            "from_cache": False,
            "fetch_date": rec2["fetch_date"] if rec2 else dt.date.today().isoformat(),
            "next_update": rec2["next_update"] if rec2 else "",
            "data": data,
        })
    except Exception as exc:
        logger.error("market-info GET error: %s", exc)
        msg = _market_info_error_message(exc)
        try:
            rec = _mi.get_cached(country, category)
            data = _market_info_cached_data(rec)
            if data:
                return jsonify({
                    "ok": True,
                    "from_cache": True,
                    "stale": True,
                    "fetch_date": rec.get("fetch_date"),
                    "next_update": rec.get("next_update"),
                    "data": data,
                    "warning": msg,
                })
        except Exception:
            pass
        status = 429 if "rate-limited" in msg or "rate limited" in msg else 500
        return jsonify({"error": msg}), status


# ── In-memory progress state (single worker) ─────────────────────────────────
_mi_refresh_state: dict = {"running": False, "country": "", "items": [], "started_at": None}


@app.route("/api/market-info/refresh-all", methods=["POST"])
@login_required
def market_info_refresh_all():
    """Refresh all active categories for a given country, one by one, with live progress."""
    import market_info as _mi
    import threading
    global _mi_refresh_state
    body = request.get_json(force=True, silent=True) or {}
    country = body.get("country", "").strip()
    if not country:
        return jsonify({"error": "country required"}), 400
    if _mi_refresh_state.get("running"):
        return jsonify({"ok": True, "message": "Already running", "state": _mi_refresh_state})
    missing = _mi_missing_keys()
    if missing:
        return jsonify({"error": f"API keys missing: {', '.join(missing)}"}), 503

    with closing(get_db()) as db:
        categories = db.execute(
            "SELECT name FROM mi_categories WHERE is_active=1 ORDER BY sort_order, name"
        ).fetchall()

    items = [{"category": c["name"], "status": "pending", "error": ""} for c in categories]
    _mi_refresh_state = {
        "running": True,
        "country": country,
        "items": items,
        "started_at": dt.datetime.now().isoformat(timespec="seconds"),
    }

    def _run():
        global _mi_refresh_state
        def _fetch_item(item):
            item["status"] = "fetching"
            try:
                _cat_meta = _get_category_meta(item["category"])
                data = _mi.fetch_market_info(country, item["category"], _cat_meta["hs_codes"], _cat_meta["products"], fast=True)
                _mi.save_cache(country, item["category"], data, triggered_by="manual")
                item["status"] = "done"
            except Exception as exc:
                cached = _mi.get_cached(country, item["category"])
                data = _market_info_cached_data(cached)
                msg = _market_info_error_message(exc)
                if data:
                    item["status"] = "done"
                    item["from_cache"] = True
                    item["warning"] = msg
                    logger.warning("refresh-all %s/%s using cached data after error: %s", country, item["category"], exc)
                else:
                    item["status"] = "error"
                    item["error"] = msg
                    logger.warning("refresh-all %s/%s failed with no cache: %s", country, item["category"], exc)

        max_workers = min(4, len(_mi_refresh_state["items"])) or 1
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = [pool.submit(_fetch_item, item) for item in _mi_refresh_state["items"]]
            for _ in as_completed(futures):
                pass
        _mi_refresh_state["running"] = False

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "state": _mi_refresh_state})


@app.route("/api/market-info/refresh-all/status", methods=["GET"])
@login_required
def market_info_refresh_all_status():
    return jsonify(_mi_refresh_state)


# ── Region refresh — fetches every country in a region for one category ────────
_mi_region_state: dict = {"running": False, "region": "", "category": "", "items": [], "started_at": None}


@app.route("/api/market-info/region-refresh", methods=["POST"])
@login_required
def market_info_region_refresh():
    """Refresh (or read fresh cache for) every country in a region, for one category,
    one by one, with live progress. Reuses each country's existing cache if fresh."""
    import market_info as _mi
    import threading
    global _mi_region_state
    body = request.get_json(force=True, silent=True) or {}
    region   = body.get("region", "").strip()
    category = body.get("category", "").strip()
    if not region or not category:
        return jsonify({"error": "region and category required"}), 400
    if _mi_region_state.get("running"):
        return jsonify({"ok": True, "message": "Already running", "state": _mi_region_state})
    missing = _mi_missing_keys()
    if missing:
        return jsonify({"error": f"API keys missing: {', '.join(missing)}"}), 503

    with closing(get_db()) as db:
        countries = [r["market"] for r in db.execute(
            "SELECT DISTINCT market FROM mi_team_markets "
            "WHERE is_active=1 AND region=? ORDER BY market", (region,)
        ).fetchall()]

    if not countries:
        return jsonify({"error": f"No countries mapped to region '{region}'"}), 400

    items = [{"country": c, "status": "pending", "error": ""} for c in countries]
    _mi_region_state = {
        "running": True,
        "region": region,
        "category": category,
        "items": items,
        "started_at": dt.datetime.now().isoformat(timespec="seconds"),
    }

    def _run():
        global _mi_region_state
        _cat_meta = _get_category_meta(category)
        def _fetch_item(item):
            item["status"] = "fetching"
            try:
                # Reuse fresh cache if available, else fetch fresh
                cached = _mi.get_cached(item["country"], category)
                if cached and _mi.is_fresh(cached) and cached.get("cache_status") == "ok":
                    item["status"] = "done"
                    item["from_cache"] = True
                    return
                data = _mi.fetch_market_info(item["country"], category, _cat_meta["hs_codes"], _cat_meta["products"], fast=True)
                _mi.save_cache(item["country"], category, data, triggered_by="region")
                item["status"] = "done"
            except Exception as exc:
                cached = _mi.get_cached(item["country"], category)
                if cached:
                    item["status"] = "done"
                    item["from_cache"] = True
                    item["warning"] = str(exc)[:200]
                    logger.warning("region-refresh %s/%s/%s using cached data after error: %s", region, item["country"], category, exc)
                else:
                    item["status"] = "done"
                    item["warning"] = "refresh deferred due to API throttling"
                    logger.warning("region-refresh %s/%s/%s deferred: %s", region, item["country"], category, exc)
        max_workers = min(4, len(_mi_region_state["items"])) or 1
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = [pool.submit(_fetch_item, item) for item in _mi_region_state["items"]]
            for _ in as_completed(futures):
                pass
        _mi_region_state["running"] = False

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "state": _mi_region_state})


@app.route("/api/market-info/region-refresh/status", methods=["GET"])
@login_required
def market_info_region_refresh_status():
    return jsonify(_mi_region_state)


@app.route("/api/market-info/region", methods=["GET"])
@login_required
def market_info_region_get():
    """Return aggregated totals for a region (sum across its countries' cached data
    for the given category) plus the per-country raw data for drill-down."""
    region   = request.args.get("region", "").strip()
    category = request.args.get("category", "").strip()
    if not region or not category:
        return jsonify({"error": "region and category required"}), 400
    try:
        import market_info as _mi
        with closing(get_db()) as db:
            countries = [r["market"] for r in db.execute(
                "SELECT DISTINCT market FROM mi_team_markets "
                "WHERE is_active=1 AND region=? ORDER BY market", (region,)
            ).fetchall()]
        if not countries:
            return jsonify({"error": f"No countries mapped to region '{region}'"}), 400

        per_country = []
        for c in countries:
            rec = _mi.get_cached(c, category)
            if rec:
                data = _market_info_cached_data(rec)
                per_country.append({"country": c, "fetch_date": rec["fetch_date"], "data": data})
            else:
                per_country.append({"country": c, "fetch_date": None, "data": None})

        # Aggregate market_size across countries that have data
        tot_total = tot_dom = tot_imp = 0.0
        have_any = False
        for pc in per_country:
            ms = (pc["data"] or {}).get("market_size") or {}
            if ms.get("total_usd_mn") is not None:
                have_any = True
                tot_total += ms.get("total_usd_mn") or 0
                tot_dom   += ms.get("domestic_usd_mn") or 0
                tot_imp   += ms.get("imports_usd_mn") or 0

        aggregate = None
        if have_any:
            aggregate = {
                "total_usd_mn": round(tot_total, 2),
                "domestic_usd_mn": round(tot_dom, 2),
                "imports_usd_mn": round(tot_imp, 2),
                "domestic_pct": round(tot_dom / tot_total * 100, 1) if tot_total > 0 else None,
                "imports_pct": round(tot_imp / tot_total * 100, 1) if tot_total > 0 else None,
                "countries_with_data": sum(1 for pc in per_country if pc["data"]),
                "countries_total": len(countries),
            }

        return jsonify({"ok": True, "region": region, "category": category,
                        "aggregate": aggregate, "countries": per_country})
    except Exception as exc:
        logger.error("market-info region GET error: %s", exc)
        return jsonify({"error": str(exc)}), 500


@app.route("/api/market-info/schedule", methods=["POST"])
@login_required
def market_info_schedule_job():
    """Queue a one-shot background refresh for the given country (or region)
    + category, to run once at the specified date/time."""
    body = request.get_json(force=True, silent=True) or {}
    country  = body.get("country", "").strip()
    region   = body.get("region", "").strip()
    category = body.get("category", "").strip()
    run_at_s = body.get("run_at", "").strip()
    if not category or not run_at_s or (not country and not region):
        return jsonify({"error": "category, run_at, and country or region are required"}), 400
    try:
        run_at = dt.datetime.fromisoformat(run_at_s)
    except ValueError:
        return jsonify({"error": "run_at must be an ISO datetime, e.g. 2026-06-25T14:30"}), 400
    if run_at <= dt.datetime.now():
        return jsonify({"error": "run_at must be in the future"}), 400
    missing = _mi_missing_keys()
    if missing:
        return jsonify({"error": f"API keys missing: {', '.join(missing)}"}), 503

    u = get_current_user()
    now = dt.datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as db:
        db.execute(
            "INSERT INTO mi_oneshot_jobs (country,region,category,run_at,status,created_by,created_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (country, region, category, run_at.isoformat(timespec="seconds"), "pending",
             u["username"] if u else "", now)
        )
        db.commit()
        job_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]

    _schedule_one_shot_mi_job(job_id, run_at)
    return jsonify({"ok": True, "id": job_id, "run_at": run_at.isoformat(timespec="seconds")}), 201


@app.route("/api/market-info/schedule", methods=["GET"])
@login_required
def market_info_schedule_list():
    """List queued/recent one-shot Market Info refresh jobs."""
    with closing(get_db()) as db:
        rows = db.execute(
            "SELECT * FROM mi_oneshot_jobs ORDER BY id DESC LIMIT 50"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/market-info/schedule/<int:job_id>", methods=["DELETE"])
@login_required
def market_info_schedule_cancel(job_id):
    """Cancel a pending one-shot job."""
    with closing(get_db()) as db:
        job = db.execute("SELECT status FROM mi_oneshot_jobs WHERE id=?", (job_id,)).fetchone()
        if not job:
            return jsonify({"error": "not found"}), 404
        if job["status"] != "pending":
            return jsonify({"error": f"cannot cancel a job with status '{job['status']}'"}), 400
        db.execute("UPDATE mi_oneshot_jobs SET status='cancelled' WHERE id=?", (job_id,))
        db.commit()
    if _scheduler is not None:
        try:
            _scheduler.remove_job(f"mi_oneshot_{job_id}")
        except Exception:
            pass
    return jsonify({"ok": True})


@app.route("/api/market-info/refresh", methods=["POST"])
@login_required
def market_info_refresh():
    body     = request.get_json(force=True, silent=True) or {}
    country  = body.get("country", "").strip()
    category = body.get("category", "").strip()
    if not country or not category:
        return jsonify({"error": "country and category required"}), 400
    try:
        import market_info as _mi
        missing = _mi_missing_keys()
        if missing:
            rec = _mi.get_cached(country, category)
            data = _market_info_cached_data(rec)
            if data:
                return jsonify({
                    "ok": True,
                    "from_cache": True,
                    "stale": True,
                    "fetch_date": rec.get("fetch_date"),
                    "next_update": rec.get("next_update"),
                    "data": data,
                    "warning": f"API keys missing: {', '.join(missing)}",
                })
            return jsonify({"error": f"API keys missing: {', '.join(missing)}"}), 503
        _cat_meta = _get_category_meta(category)
        data = _mi.fetch_market_info(country, category, _cat_meta["hs_codes"], _cat_meta["products"], fast=True)
        _mi.save_cache(country, category, data, triggered_by="manual")
        rec = _mi.get_cached(country, category)
        return jsonify({
            "ok": True,
            "fetch_date": rec["fetch_date"] if rec else dt.date.today().isoformat(),
            "next_update": rec["next_update"] if rec else "",
            "data": data,
        })
    except Exception as exc:
        logger.error("market-info refresh error: %s", exc)
        msg = _market_info_error_message(exc)
        try:
            rec = _mi.get_cached(country, category)
            data = _market_info_cached_data(rec)
            if data:
                return jsonify({
                    "ok": True,
                    "from_cache": True,
                    "stale": True,
                    "fetch_date": rec.get("fetch_date"),
                    "next_update": rec.get("next_update"),
                    "data": data,
                    "warning": msg,
                })
        except Exception:
            pass
        status = 429 if "rate-limited" in msg or "rate limited" in msg else 500
        return jsonify({"error": msg}), status


@app.route("/api/market-info/admin-alerts", methods=["GET"])
@login_required
def market_info_admin_alerts():
    """Return cache entries that have hit max attempts."""
    try:
        import market_info as _mi
        with closing(get_db()) as db:
            rows = db.execute(
                "SELECT country, category, attempt_count, fetch_date, cache_status FROM market_info_cache WHERE attempt_count >= ?",
                (_mi.MAX_ATTEMPTS,)
            ).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ── Serve frontend ──
@app.route("/")
def index():
    return send_from_directory(FRONTEND_DIR, "index.html")


@app.route("/<path:path>")
def static_proxy(path):
    return send_from_directory(FRONTEND_DIR, path)


init_db()
_seed_mi_defaults()


_scheduler = None  # global BackgroundScheduler instance, set by _start_scheduler()


def _run_market_info_refresh():
    try:
        import market_info as _minfo
        def _alert(country, category, attempts):
            logger.error(
                "ADMIN ALERT: market info for %s/%s failed %d times — manual intervention needed",
                country, category, attempts,
            )
        _minfo.refresh_all_stale(alert_callback=_alert)
    except Exception as exc:
        logger.error("Market info scheduler error: %s", exc)


def _run_one_shot_mi_job(job_id):
    """Execute a single queued one-shot Market Info refresh job (see mi_oneshot_jobs)."""
    import market_info as _minfo
    with closing(get_db()) as db:
        job = db.execute("SELECT * FROM mi_oneshot_jobs WHERE id=?", (job_id,)).fetchone()
        if not job or job["status"] != "pending":
            return
        db.execute("UPDATE mi_oneshot_jobs SET status='running' WHERE id=?", (job_id,))
        db.commit()

    try:
        category = job["category"]
        _cat_meta = _get_category_meta(category)
        if job["region"]:
            with closing(get_db()) as db:
                countries = [r["market"] for r in db.execute(
                    "SELECT DISTINCT market FROM mi_team_markets WHERE is_active=1 AND region=?",
                    (job["region"],)
                ).fetchall()]
            for c in countries:
                data = _minfo.fetch_market_info(c, category, _cat_meta["hs_codes"], _cat_meta["products"], fast=True)
                _minfo.save_cache(c, category, data, triggered_by="scheduled")
        else:
            data = _minfo.fetch_market_info(job["country"], category, _cat_meta["hs_codes"], _cat_meta["products"], fast=True)
            _minfo.save_cache(job["country"], category, data, triggered_by="scheduled")
        finished_at = dt.datetime.now().isoformat(timespec="seconds")
        with closing(get_db()) as db:
            db.execute("UPDATE mi_oneshot_jobs SET status='done', completed_at=? WHERE id=?", (finished_at, job_id))
            db.commit()
    except Exception as exc:
        logger.error("One-shot MI job %d failed: %s", job_id, exc)
        finished_at = dt.datetime.now().isoformat(timespec="seconds")
        with closing(get_db()) as db:
            db.execute(
                "UPDATE mi_oneshot_jobs SET status='error', error=?, completed_at=? WHERE id=?",
                (str(exc)[:300], finished_at, job_id)
            )
            db.commit()


def _schedule_one_shot_mi_job(job_id, run_at):
    """Register a one-shot APScheduler DateTrigger job for the given mi_oneshot_jobs row."""
    if _scheduler is None:
        return
    from apscheduler.triggers.date import DateTrigger
    _scheduler.add_job(
        _run_one_shot_mi_job,
        DateTrigger(run_date=run_at),
        args=[job_id],
        id=f"mi_oneshot_{job_id}",
        replace_existing=True,
        misfire_grace_time=3600,
    )


def _start_scheduler():
    """Start APScheduler for the monthly intelligence job and the daily Market
    Info stale-cache refresh. Only runs in the main gunicorn worker process
    (single worker setup). For multi-worker deployments use a DB-backed
    scheduler (SQLAlchemyJobStore). To change either schedule, edit the
    CronTrigger arguments below.
    One-shot, user-scheduled Market Info jobs (see mi_oneshot_jobs /
    /api/market-info/schedule) are registered separately via
    _schedule_one_shot_mi_job() and are re-armed for any still-pending jobs
    found in the DB at startup (so they survive a container restart)."""
    global _scheduler
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from apscheduler.triggers.cron import CronTrigger
        import intelligence as _mi

        sched = BackgroundScheduler(daemon=True)
        # Runs at 02:00 on the 1st of every month.
        sched.add_job(
            _mi.run_intelligence_job,
            CronTrigger(day=1, hour=2, minute=0),
            id="monthly_intelligence",
            replace_existing=True,
            misfire_grace_time=3600,
        )
        # Market info: refresh stale entries daily at 03:00
        sched.add_job(
            _run_market_info_refresh,
            CronTrigger(hour=3, minute=0),
            id="daily_market_info_refresh",
            replace_existing=True,
            misfire_grace_time=3600,
        )
        sched.start()
        _scheduler = sched

        # Re-arm any pending one-shot jobs that didn't fire before a restart
        with closing(get_db()) as db:
            pending = db.execute(
                "SELECT id, run_at FROM mi_oneshot_jobs WHERE status='pending'"
            ).fetchall()
        for p in pending:
            _schedule_one_shot_mi_job(p["id"], dt.datetime.fromisoformat(p["run_at"]))

        logger.info("Schedulers started: monthly intelligence (1st@02:00), daily market info refresh (03:00), %d pending one-shot job(s) re-armed", len(pending))
    except ImportError as exc:
        logger.warning("APScheduler or intelligence module not available — scheduler not started: %s", exc)
    except Exception as exc:
        logger.error("Scheduler startup error: %s", exc)


_start_scheduler()

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8000, debug=False, use_reloader=False)
